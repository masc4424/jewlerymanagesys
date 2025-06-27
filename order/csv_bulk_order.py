import csv
import json
from datetime import datetime
import pandas as pd
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.contrib import messages
from django.db import transaction
from product_inv.models import *
from order.models import *
from user_role_management.models import *
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import io


@login_required
def download_order_template(request):
    """Generate and download Excel template for bulk order upload with comprehensive client-model mapping"""
    # Get filter parameters
    client_username = request.GET.get('client_username')
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Order Template"
    
    # Create data sheet for dropdowns
    data_ws = wb.create_sheet("Data")
    
    # Create client-model mapping sheet with separate rows
    mapping_ws = wb.create_sheet("Client_Model_Mapping")
    
    # Write headers with styling
    headers = [
        'Client Username',
        'Model Number',
        'Color Name',
        'Quantity',
        'Estimated Delivery Date (YYYY-MM-DD)',
        'Status (optional)'
    ]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    
    # Get users with role = 'client' using the UserRole model
    try:
        client_role = Role.objects.get(role_unique_id='client')
        client_users = User.objects.filter(userrole__role=client_role).distinct().order_by('first_name', 'last_name', 'username')
        
        # Format as "first_name last_name : username"
        all_clients = []
        client_username_map = {}  # To map display format back to username
        
        for user in client_users:
            display_name = f"{user.first_name} {user.last_name} : {user.username}".strip()
            # Handle cases where first_name or last_name might be empty
            if not user.first_name and not user.last_name:
                display_name = f" : {user.username}"
            elif not user.first_name:
                display_name = f"{user.last_name} : {user.username}"
            elif not user.last_name:
                display_name = f"{user.first_name} : {user.username}"
            
            all_clients.append(display_name)
            client_username_map[display_name] = user.username
            
    except Role.DoesNotExist:
        print("ERROR: Role with role_unique_id='client' does not exist")
        all_clients = []
        client_username_map = {}
    except Exception as e:
        print(f"ERROR fetching clients with role='client': {str(e)}")
        # Fallback to all users with same format
        all_users = User.objects.all().distinct().order_by('first_name', 'last_name', 'username')
        all_clients = []
        client_username_map = {}
        
        for user in all_users:
            display_name = f"{user.first_name} {user.last_name} : {user.username}".strip()
            if not user.first_name and not user.last_name:
                display_name = f" : {user.username}"
            elif not user.first_name:
                display_name = f"{user.last_name} : {user.username}"
            elif not user.last_name:
                display_name = f"{user.first_name} : {user.username}"
            
            all_clients.append(display_name)
            client_username_map[display_name] = user.username
    
    # Debug information
    print(f"=== CLIENT LOADING DEBUG ===")
    print(f"Total users in system: {User.objects.count()}")
    print(f"Total clients loaded: {len(all_clients)}")
    print(f"First 5 clients: {all_clients[:5]}")
    
    # Get ALL models from the system
    all_models = list(
        Model.objects.all()
        .values_list('model_no', flat=True)
        .distinct()
        .order_by('model_no')
    )
    
    # Build comprehensive client-model mapping for reference
    client_model_mapping = {}
    model_color_map = {}
    
    # Get all client-model relationships for mapping reference
    try:
        client_role = Role.objects.get(role_unique_id='client')
        model_client_relationships = ModelClient.objects.select_related('model', 'client').filter(
            client__userrole__role=client_role
        )
    except Role.DoesNotExist:
        model_client_relationships = ModelClient.objects.select_related('model', 'client').all()
    
    for mc_rel in model_client_relationships:
        # Create display name for client
        user = mc_rel.client
        client_display = f"{user.first_name} {user.last_name} : {user.username}".strip()
        if not user.first_name and not user.last_name:
            client_display = f" : {user.username}"
        elif not user.first_name:
            client_display = f"{user.last_name} : {user.username}"  
        elif not user.last_name:
            client_display = f"{user.first_name} : {user.username}"
            
        model_no = mc_rel.model.model_no
        
        if client_display not in client_model_mapping:
            client_model_mapping[client_display] = []
        if model_no not in client_model_mapping[client_display]:
            client_model_mapping[client_display].append(model_no)
    
    # Build model-color mapping for ALL models
    for model in Model.objects.all():
        model_colors = list(
            ModelColor.objects.filter(model=model)
            .values_list('color', flat=True)
            .order_by('color')
        )
        model_color_map[model.model_no] = model_colors
    
    # Sort models for each client
    for client_display in client_model_mapping:
        client_model_mapping[client_display].sort()
    
    # Get all colors and statuses
    all_colors = list(
        ModelColor.objects.values_list('color', flat=True)
        .distinct()
        .order_by('color')
    )
    
    all_statuses = list(
        ModelStatus.objects.values_list('status', flat=True)
        .order_by('status')
    )
    
    # === POPULATE DATA SHEETS ===
    
    # 1. All clients list (Column A in Data sheet)
    if all_clients:
        data_ws.cell(row=1, column=1, value='All_Clients')
        for i, client in enumerate(all_clients, 2):
            data_ws.cell(row=i, column=1, value=client)
    
    # 2. All models list (Column B in Data sheet) 
    if all_models:
        data_ws.cell(row=1, column=2, value='All_Models')
        for i, model in enumerate(all_models, 2):
            data_ws.cell(row=i, column=2, value=model)
    
    # 3. All colors list (Column C in Data sheet)
    if all_colors:
        data_ws.cell(row=1, column=3, value='All_Colors')
        for i, color in enumerate(all_colors, 2):
            data_ws.cell(row=i, column=3, value=color)
    
    # 4. Status list (Column D in Data sheet)
    if all_statuses:
        data_ws.cell(row=1, column=4, value='All_Statuses')
        for i, status in enumerate(all_statuses, 2):
            data_ws.cell(row=i, column=4, value=status)
    
    # === IMPROVED CLIENT-MODEL MAPPING SHEET ===
    # Create headers for mapping sheet
    mapping_headers = ['Client Username', 'Model Number', 'Available Colors', 'Color Count']
    
    # Style mapping headers
    for col_num, header in enumerate(mapping_headers, 1):
        cell = mapping_ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Populate mapping data - SEPARATE ROW FOR EACH CLIENT-MODEL COMBINATION
    mapping_row = 2
    
    for client_display in sorted(client_model_mapping.keys()):
        models = client_model_mapping[client_display]
        
        for model_no in sorted(models):
            # Get colors for this model
            colors = model_color_map.get(model_no, [])
            
            # Write client display name
            mapping_ws.cell(row=mapping_row, column=1, value=client_display)
            
            # Write model number
            mapping_ws.cell(row=mapping_row, column=2, value=model_no)
            
            # Write colors (grouped and formatted)
            if colors:
                colors_str = ', '.join(sorted(colors))
                mapping_ws.cell(row=mapping_row, column=3, value=colors_str)
                mapping_ws.cell(row=mapping_row, column=4, value=len(colors))
            else:
                mapping_ws.cell(row=mapping_row, column=3, value='No colors available')
                mapping_ws.cell(row=mapping_row, column=4, value=0)
            
            # Style the cells
            for col in range(1, 5):
                cell = mapping_ws.cell(row=mapping_row, column=col)
                if mapping_row % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            mapping_row += 1
    
    # Add totals row
    if mapping_row > 2:
        total_row = mapping_row + 1
        mapping_ws.cell(row=total_row, column=1, value="TOTAL COMBINATIONS:")
        mapping_ws.cell(row=total_row, column=2, value=mapping_row - 2)
        
        # Style total row
        for col in range(1, 3):
            cell = mapping_ws.cell(row=total_row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # === ADD DATA VALIDATION WITH ERROR HANDLING ===
    
    try:
        # Client dropdown (column A)
        if all_clients:
            client_range = f"Data!A2:A{len(all_clients)+1}"
            client_dv = DataValidation(
                type="list",
                formula1=client_range,
                allow_blank=False
            )
            client_dv.error = "Please select a valid client from the list"
            client_dv.errorTitle = "Invalid Client"
            ws.add_data_validation(client_dv)
            client_dv.add('A2:A1000')
        
        # Model dropdown (column B)
        if all_models:
            model_range = f"Data!B2:B{len(all_models)+1}"
            model_dv = DataValidation(
                type="list", 
                formula1=model_range,
                allow_blank=False
            )
            model_dv.error = "Please select a model from the list"
            model_dv.errorTitle = "Invalid Model"
            ws.add_data_validation(model_dv)
            model_dv.add('B2:B1000')
        
        # Color dropdown (column C)
        if all_colors:
            color_range = f"Data!C2:C{len(all_colors)+1}"
            color_dv = DataValidation(
                type="list",
                formula1=color_range,
                allow_blank=True
            )
            color_dv.error = "Please select a valid color from the list"
            color_dv.errorTitle = "Invalid Color"
            ws.add_data_validation(color_dv)
            color_dv.add('C2:C1000')
        
        # Status dropdown (column F)
        if all_statuses:
            status_range = f"Data!D2:D{len(all_statuses)+1}"
            status_dv = DataValidation(
                type="list",
                formula1=status_range,
                allow_blank=True
            )
            status_dv.error = "Please select a valid status from the list"
            status_dv.errorTitle = "Invalid Status"
            ws.add_data_validation(status_dv)
            status_dv.add('F2:F1000')
            
    except Exception as e:
        print(f"Error adding data validation: {str(e)}")
    
    # === ADD HELPER INFORMATION ===
    helper_data = [
        '[Select Client: FirstName LastName : username]',
        '[Select Model]',
        '[Select Color]',
        '[Enter Quantity]',
        '[YYYY-MM-DD]',
        '[Optional Status]'
    ]
    
    for col, helper_text in enumerate(helper_data, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = helper_text
        cell.font = Font(italic=True, color="666666")
        cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    # === AUTO-ADJUST COLUMN WIDTHS ===
    def adjust_column_widths(worksheet):
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    continue
            
            # Set reasonable width limits
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Apply width adjustments
    adjust_column_widths(ws)
    adjust_column_widths(mapping_ws)
    
    # === ENHANCED INSTRUCTIONS SHEET ===
    instructions_ws = wb.create_sheet("Instructions")
    instructions = [
        "EXCEL ORDER UPLOAD TEMPLATE - INSTRUCTIONS",
        "=" * 50,
        "",
        "IMPORTANT CHANGES:",
        "• Client format: 'FirstName LastName : username'",
        "• Client-Model mapping shows SEPARATE ROWS for each combination",
        "• Colors are grouped and counted for each model",
        "• Better data validation and error handling",
        "",
        "HOW TO USE:",
        "",
        "1. CHECK MAPPING SHEET:",
        "   • Open 'Client_Model_Mapping' sheet",
        "   • Each row shows one client-model combination",
        "   • Client format: 'FirstName LastName : username'",
        "   • Colors are listed and counted for each model",
        "",
        "2. FILL ORDER DATA (starting from row 3):",
        "   • Column A: Select client (FirstName LastName : username format)",
        "   • Column B: Select model (check mapping first!)",
        "   • Column C: Select color for the model",
        "   • Column D: Enter quantity (positive number)",
        "   • Column E: Enter date as YYYY-MM-DD",
        "   • Column F: Select status (optional)",
        "",
        "3. VALIDATION:",
        "   • All dropdowns are validated",
        "   • Check mapping sheet for valid combinations",
        "   • Server will validate final upload",
        "",
        f"4. CURRENT DATA SUMMARY:",
        f"   • Total clients: {len(all_clients)}",
        f"   • Total models: {len(all_models)}",
        f"   • Total colors: {len(all_colors)}",
        f"   • Client-model combinations: {sum(len(models) for models in client_model_mapping.values())}",
    ]
    
    # Add client-specific info if requested
    if client_username:
        # Find the display name that matches the username
        matching_client_display = None
        for display_name, username in client_username_map.items():
            if username == client_username:
                matching_client_display = display_name
                break
        
        if matching_client_display and matching_client_display in client_model_mapping:
            models = client_model_mapping[matching_client_display]
            instructions.extend([
                "",
                f"5. INFO FOR CLIENT '{matching_client_display}':",
                f"   • You have {len(models)} authorized models",
                "   • Check mapping sheet for color details",
            ])
        else:
            instructions.extend([
                "",
                f"5. WARNING FOR CLIENT '{client_username}':",
                "   • No authorized models found for this client",
                "   • Please contact administrator",
            ])
    
    # Write instructions
    for i, instruction in enumerate(instructions, 1):
        cell = instructions_ws.cell(row=i, column=1)
        cell.value = instruction
        
        # Style instructions
        if instruction.startswith('EXCEL ORDER') or instruction.startswith('IMPORTANT'):
            cell.font = Font(bold=True, size=14, color="C00000")
        elif instruction.endswith(':') and not instruction.startswith('   '):
            cell.font = Font(bold=True, color="0066CC")
        elif instruction.startswith('   • '):
            cell.font = Font(color="333333")
    
    adjust_column_widths(instructions_ws)
    
    # Hide data sheet
    data_ws.sheet_state = 'hidden'
    
    # === SAVE WITH PROPER ERROR HANDLING ===
    try:
        # Create filename FIRST
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if client_username:
            # Find matching display name for filename
            matching_client_display = None
            for display_name, username in client_username_map.items():
                if username == client_username:
                    matching_client_display = display_name
                    break
            
            if matching_client_display and matching_client_display in client_model_mapping:
                model_count = len(client_model_mapping[matching_client_display])
                # Clean filename by removing/replacing invalid characters
                clean_client = re.sub(r'[^\w\-_.]', '_', client_username)
                filename = f"order_template_{clean_client}_{model_count}models_{timestamp}.xlsx"
            else:
                clean_client = re.sub(r'[^\w\-_.]', '_', client_username)
                filename = f"order_template_{clean_client}_nomodels_{timestamp}.xlsx"
        else:
            filename = f"order_template_{len(all_clients)}clients_{timestamp}.xlsx"
        
        # Create BytesIO buffer
        buffer = io.BytesIO()
        
        # Save workbook to buffer
        wb.save(buffer)
        buffer.seek(0)
        
        # Create HTTP response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Set headers for file download
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = len(buffer.getvalue())
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        
        # Close buffer
        buffer.close()
        
        print(f"Successfully generated Excel file: {filename}")
        
        return response
        
    except Exception as e:
        print(f"Error generating Excel file: {str(e)}")
        import traceback
        traceback.print_exc()
        return HttpResponse(f"Error generating Excel file: {str(e)}", status=500)

@login_required
def get_template_data(request):
    """API endpoint to get dropdown data for frontend with cascading filters"""
    from .models import ModelClient, ModelColor, ModelStatus
    
    # Get filter parameters
    jewelry_type = request.GET.get('jewelry_type')
    model_no = request.GET.get('model_no')
    client_username = request.GET.get('client_username')
    
    # Start with all ModelClient relationships
    queryset = ModelClient.objects.select_related('model', 'client')
    
    # Get all jewelry types
    jewelry_types = list(
        queryset.values_list('model__jewelry_type', flat=True)
        .distinct()
        .order_by('model__jewelry_type')
    )
    jewelry_types = [jt for jt in jewelry_types if jt]  # Remove None values
    
    # Apply jewelry type filter if provided
    if jewelry_type:
        queryset = queryset.filter(model__jewelry_type=jewelry_type)
    
    # Get models based on current filter
    models = list(
        queryset.values_list('model__model_no', flat=True)
        .distinct()
        .order_by('model__model_no')
    )
    
    # Apply model filter if provided
    if model_no:
        queryset = queryset.filter(model__model_no=model_no)
    
    # Get clients based on current filter
    clients = list(
        queryset.values_list('client__username', flat=True)
        .distinct()
        .order_by('client__username')
    )
    
    # Apply client filter if provided
    if client_username:
        queryset = queryset.filter(client__username=client_username)
    
    # Build client-model mapping based on current filters
    client_models = {}
    model_colors_map = {}
    model_jewelry_types = {}
    
    for mc in queryset:
        client_user = mc.client.username
        model_number = mc.model.model_no
        
        if client_user not in client_models:
            client_models[client_user] = []
        if model_number not in client_models[client_user]:
            client_models[client_user].append(model_number)
        
        # Get colors for this specific model
        if model_number not in model_colors_map:
            model_colors = list(ModelColor.objects.filter(model=mc.model).values_list('color', flat=True))
            model_colors_map[model_number] = model_colors
        
        # Map model to jewelry type
        if model_number not in model_jewelry_types:
            model_jewelry_types[model_number] = getattr(mc.model, 'jewelry_type', '')
    
    # Get all statuses
    statuses = list(ModelStatus.objects.values_list('status', flat=True).order_by('status'))
    
    return JsonResponse({
        'jewelry_types': jewelry_types,
        'models': models,
        'clients': clients,
        'client_models': client_models,
        'model_colors': model_colors_map,
        'model_jewelry_types': model_jewelry_types,
        'statuses': statuses,
        'applied_filters': {
            'jewelry_type': jewelry_type,
            'model_no': model_no,
            'client_username': client_username
        }
    })


@login_required
def get_model_colors(request, model_no):
    """API endpoint to get colors for a specific model"""
    try:
        model = Model.objects.get(model_no=model_no)
        colors = list(ModelColor.objects.filter(model=model).values_list('color', flat=True))
        return JsonResponse({'colors': colors})
    except Model.DoesNotExist:
        return JsonResponse({'error': 'Model not found'}, status=404)


@login_required
def get_jewelry_types(request):
    """API endpoint to get all available jewelry types"""
    from .models import ModelClient
    
    jewelry_types = list(
        ModelClient.objects.select_related('model')
        .values_list('model__jewelry_type', flat=True)
        .distinct()
        .order_by('model__jewelry_type')
    )
    jewelry_types = [jt for jt in jewelry_types if jt]  # Remove None values
    
    return JsonResponse({'jewelry_types': jewelry_types})


@login_required
def get_models_by_jewelry_type(request, jewelry_type):
    """API endpoint to get models filtered by jewelry type"""
    from .models import ModelClient
    
    models = list(
        ModelClient.objects.select_related('model')
        .filter(model__jewelry_type=jewelry_type)
        .values_list('model__model_no', flat=True)
        .distinct()
        .order_by('model__model_no')
    )
    
    return JsonResponse({'models': models})


@login_required
def get_clients_by_filters(request):
    """API endpoint to get clients filtered by jewelry type and/or model"""
    from .models import ModelClient
    
    jewelry_type = request.GET.get('jewelry_type')
    model_no = request.GET.get('model_no')
    
    queryset = ModelClient.objects.select_related('model', 'client')
    
    if jewelry_type:
        queryset = queryset.filter(model__jewelry_type=jewelry_type)
    
    if model_no:
        queryset = queryset.filter(model__model_no=model_no)
    
    clients = list(
        queryset.values_list('client__username', flat=True)
        .distinct()
        .order_by('client__username')
    )
    
    return JsonResponse({'clients': clients})
    

@login_required
def download_existing_orders_template(request):
    """Download CSV with existing order data for reference with optional filtering"""
    # Get filter parameters
    jewelry_type = request.GET.get('jewelry_type')
    model_no = request.GET.get('model_no')
    client_username = request.GET.get('client_username')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="existing_orders_data.csv"'
    
    writer = csv.writer(response)
    
    # Write headers
    headers = [
        'Order ID',
        'Client Username',
        'Model Number',
        'Jewelry Type',
        'Color Name',
        'Quantity',
        'Quantity Delivered',
        'Date of Order',
        'Estimated Delivery Date',
        'Status',
        'Delivered',
        'Approved'
    ]
    writer.writerow(headers)
    
    # Get filtered orders
    orders = Order.objects.select_related('client', 'model', 'color', 'status')
    
    if jewelry_type:
        orders = orders.filter(model__jewelry_type=jewelry_type)
    
    if model_no:
        orders = orders.filter(model__model_no=model_no)
        
    if client_username:
        orders = orders.filter(client__username=client_username)
    
    orders = orders.all()
    
    for order in orders:
        row = [
            order.id,
            order.client.username if order.client else 'No Client',
            order.model.model_no,
            getattr(order.model, 'jewelry_type', 'N/A'),
            order.color.color_name if order.color else 'No Color',
            order.quantity,
            order.quantity_delivered,
            order.date_of_order.strftime('%Y-%m-%d'),
            order.est_delivery_date.strftime('%Y-%m-%d'),
            order.status.status_name if order.status else 'No Status',
            'Yes' if order.delivered else 'No',
            'Yes' if order.is_approved else 'No'
        ]
        writer.writerow(row)
    
    return response


@csrf_exempt
@login_required
def validate_bulk_upload(request):
    """Validate uploaded CSV/Excel file before processing"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    if 'file' not in request.FILES:
        return JsonResponse({'error': 'No file uploaded'}, status=400)
    
    file = request.FILES['file']
    
    # Validate file type
    if not file.name.endswith(('.csv', '.xlsx', '.xls')):
        return JsonResponse({'error': 'Please upload a CSV or Excel file'}, status=400)
    
    try:
        # Read file based on extension
        if file.name.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        
        errors = []
        warnings = []
        valid_rows = 0
        
        required_columns = [
            'Client Username',
            'Model Number', 
            'Quantity',
            'Estimated Delivery Date (YYYY-MM-DD)'
        ]
        
        # Check required columns
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            errors.append(f"Missing required columns: {', '.join(missing_columns)}")
            return JsonResponse({
                'valid': False,
                'errors': errors,
                'warnings': warnings,
                'valid_rows': 0,
                'total_rows': len(df)
            })
        
        # Helper function to extract username from "FirstName LastName : username" format
        def extract_username(client_value):
            """Extract username from 'FirstName LastName : username' format"""
            if pd.isna(client_value):
                return None
            
            client_str = str(client_value).strip()
            if not client_str:
                return None
            
            # Check if it contains the colon separator
            if ':' in client_str:
                # Split by colon and get the part after it (username)
                parts = client_str.split(':')
                if len(parts) >= 2:
                    username = parts[-1].strip()  # Get last part in case there are multiple colons
                    return username if username else None
                else:
                    return None
            else:
                # If no colon, assume it's just the username
                return client_str
        
        # Helper function to normalize date format
        def normalize_date(date_value):
            """Convert various date formats to YYYY-MM-DD string"""
            if pd.isna(date_value):
                return None, "Date is required"
            
            try:
                # Handle different date input types
                if isinstance(date_value, pd.Timestamp):
                    # If it's already a pandas Timestamp, extract date part
                    return date_value.strftime('%Y-%m-%d'), None
                elif isinstance(date_value, datetime):
                    # If it's a datetime object, extract date part
                    return date_value.strftime('%Y-%m-%d'), None
                else:
                    # If it's a string, clean and validate
                    date_str = str(date_value).strip()
                    
                    # Remove time part if present (e.g., "2024-01-15 00:00:00" -> "2024-01-15")
                    if ' ' in date_str:
                        date_str = date_str.split(' ')[0]
                    
                    # Handle Excel date serial numbers (common issue)
                    if date_str.replace('.', '').isdigit():
                        # Convert Excel serial date to datetime
                        excel_date = pd.to_datetime(float(date_str), unit='D', origin='1899-12-30')
                        return excel_date.strftime('%Y-%m-%d'), None
                    
                    # Try different common date formats
                    date_formats = [
                        '%Y-%m-%d',      # 2024-01-15
                        '%d/%m/%Y',      # 15/01/2024
                        '%m/%d/%Y',      # 01/15/2024
                        '%d-%m-%Y',      # 15-01-2024
                        '%m-%d-%Y',      # 01-15-2024
                    ]
                    
                    for fmt in date_formats:
                        try:
                            parsed_date = datetime.strptime(date_str, fmt)
                            return parsed_date.strftime('%Y-%m-%d'), None
                        except ValueError:
                            continue
                    
                    return None, f"Invalid date format. Use YYYY-MM-DD (got '{date_value}')"
                    
            except Exception as e:
                return None, f"Date parsing error: {str(e)}"
        
        # Validate each row
        for index, row in df.iterrows():
            row_num = index + 2  # +2 because pandas is 0-indexed and we have header
            
            # Check for empty required fields
            client_display = row['Client Username']
            if pd.isna(client_display) or str(client_display).strip() == '':
                errors.append(f"Row {row_num}: Client Username is required")
                continue
            
            # Extract actual username from display format
            username = extract_username(client_display)
            if not username:
                errors.append(f"Row {row_num}: Invalid client format. Expected 'FirstName LastName : username' but got '{client_display}'")
                continue
                
            if pd.isna(row['Model Number']) or str(row['Model Number']).strip() == '':
                errors.append(f"Row {row_num}: Model Number is required")
                continue
                
            if pd.isna(row['Quantity']):
                errors.append(f"Row {row_num}: Quantity is required")
                continue
                
            if pd.isna(row['Estimated Delivery Date (YYYY-MM-DD)']):
                errors.append(f"Row {row_num}: Estimated Delivery Date is required")
                continue
            
            # Validate client exists using extracted username
            try:
                client = User.objects.get(username=username)
            except User.DoesNotExist:
                errors.append(f"Row {row_num}: Client with username '{username}' does not exist (from '{client_display}')")
                continue
            
            # Validate model exists
            model_no = str(row['Model Number']).strip()
            try:
                model = Model.objects.get(model_no=model_no)
            except Model.DoesNotExist:
                errors.append(f"Row {row_num}: Model '{model_no}' does not exist")
                continue
            
            # Validate client-model relationship exists
            if not ModelClient.objects.filter(client=client, model=model).exists():
                errors.append(f"Row {row_num}: Client '{username}' is not authorized for model '{model.model_no}'")
                continue
            
            # Validate quantity
            try:
                qty = int(row['Quantity'])
                if qty <= 0:
                    errors.append(f"Row {row_num}: Quantity must be greater than 0")
                    continue
            except (ValueError, TypeError):
                errors.append(f"Row {row_num}: Quantity must be a valid number")
                continue
            
            # Validate and normalize date format
            normalized_date, date_error = normalize_date(row['Estimated Delivery Date (YYYY-MM-DD)'])
            if date_error:
                errors.append(f"Row {row_num}: {date_error}")
                continue
            
            # Validate color if provided
            if 'Color Name' in df.columns and not pd.isna(row['Color Name']):
                color_name = str(row['Color Name']).strip()
                if color_name:
                    if not ModelColor.objects.filter(model=model, color=color_name).exists():
                        # Get available colors for better error message
                        available_colors = list(ModelColor.objects.filter(model=model).values_list('color', flat=True))
                        if available_colors:
                            warnings.append(f"Row {row_num}: Color '{color_name}' is not available for model '{model.model_no}'. Available colors: {', '.join(available_colors)}")
                        else:
                            warnings.append(f"Row {row_num}: No colors are available for model '{model.model_no}'")
            
            # Validate status if provided
            if 'Status (optional)' in df.columns and not pd.isna(row['Status (optional)']):
                status_name = str(row['Status (optional)']).strip()
                if status_name:
                    if not ModelStatus.objects.filter(status=status_name).exists():
                        # Get available statuses for better error message
                        available_statuses = list(ModelStatus.objects.values_list('status', flat=True))
                        warnings.append(f"Row {row_num}: Status '{status_name}' does not exist. Available statuses: {', '.join(available_statuses)}")
            
            valid_rows += 1
        
        # Additional summary information
        summary_info = {
            'total_rows_processed': len(df),
            'valid_rows': valid_rows,
            'error_count': len(errors),
            'warning_count': len(warnings)
        }
        
        return JsonResponse({
            'valid': len(errors) == 0,
            'errors': errors,
            'warnings': warnings,
            'valid_rows': valid_rows,
            'total_rows': len(df),
            'summary': summary_info
        })
        
    except pd.errors.EmptyDataError:
        return JsonResponse({'error': 'The uploaded file is empty'}, status=400)
    except pd.errors.ParserError as e:
        return JsonResponse({'error': f'Error parsing file: {str(e)}'}, status=400)
    except Exception as e:
        import traceback
        print(f"Validation error: {str(e)}")
        traceback.print_exc()
        return JsonResponse({'error': f'Error processing file: {str(e)}'}, status=400)
    
    
@csrf_exempt
@login_required
def process_bulk_upload(request):
    """Process validated CSV/Excel file and create orders"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    if 'file' not in request.FILES:
        return JsonResponse({'error': 'No file uploaded'}, status=400)
    
    file = request.FILES['file']
    
    try:
        # Read file
        if file.name.endswith('.csv'):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        
        created_orders = []
        failed_orders = []
        
        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    # Extract username from "Full Name : username" format
                    client_field = str(row['Client Username']).strip()
                    if ':' in client_field:
                        username = client_field.split(':')[1].strip()
                    else:
                        username = client_field
                    
                    # Get client
                    client = User.objects.get(username=username)
                    
                    # Get model
                    model = Model.objects.get(model_no=str(row['Model Number']).strip())
                    
                    # Validate client-model relationship
                    if not ModelClient.objects.filter(client=client, model=model).exists():
                        failed_orders.append({
                            'row': index + 2,
                            'error': f'Client {client.username} is not authorized for model {model.model_no}'
                        })
                        continue
                    
                    # Get color (optional)
                    color = None
                    if 'Color Name' in df.columns and not pd.isna(row['Color Name']):
                        color_name = str(row['Color Name']).strip()
                        if color_name:
                            try:
                                color = ModelColor.objects.get(model=model, color=color_name)
                            except ModelColor.DoesNotExist:
                                pass
                    
                    # Get status (optional)
                    status = None
                    if 'Status (optional)' in df.columns and not pd.isna(row['Status (optional)']):
                        status_name = str(row['Status (optional)']).strip()
                        if status_name:
                            try:
                                status = ModelStatus.objects.get(status=status_name)
                            except ModelStatus.DoesNotExist:
                                pass
                    
                    # Parse delivery date - handle both date and datetime formats
                    date_str = str(row['Estimated Delivery Date (YYYY-MM-DD)']).strip()
                    
                    # Remove time component if present
                    if ' ' in date_str:
                        date_str = date_str.split(' ')[0]
                    
                    # Create order
                    order = Order.objects.create(
                        client=client,
                        model=model,
                        color=color,
                        status=status,
                        quantity=int(row['Quantity']),
                        est_delivery_date=datetime.strptime(date_str, '%Y-%m-%d').date()
                    )
                    
                    # Get jewelry type as string
                    jewelry_type = getattr(model, 'jewelry_type', None)
                    jewelry_type_str = str(jewelry_type) if jewelry_type else 'N/A'
                    
                    created_orders.append({
                        'id': order.id,
                        'client': client.username,
                        'model': model.model_no,
                        'jewelry_type': jewelry_type_str,
                        'quantity': order.quantity
                    })
                    
                except Exception as e:
                    failed_orders.append({
                        'row': index + 2,
                        'error': str(e)
                    })
        
        return JsonResponse({
            'success': True,
            'created_count': len(created_orders),
            'failed_count': len(failed_orders),
            'created_orders': created_orders,
            'failed_orders': failed_orders
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error processing upload: {str(e)}'}, status=500)
      