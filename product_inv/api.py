from django.http import JsonResponse
from jewl_stones.models import Stone, StoneType, StoneTypeDetail
from jewl_metals.models import *
from user_role_management.models import *
from product_inv.models import *
from django.db.models import Sum
from django.http import JsonResponse, HttpResponse
from decimal import Decimal
from django.db.models import Count
from django.shortcuts import render, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
import os
from django.conf import settings
import json
from django.utils import timezone
import datetime
from django.db.models.functions import Abs
from django.db.models import F
import uuid
from django.db.models import OuterRef, Subquery
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
import io
import pandas as pd

def get_model_distribution(model_no):
    try:
        model = Model.objects.get(model_no=model_no)
        
        # STONE SECTION
        # Get all stones used in this model via RawStones
        raw_stones = RawStones.objects.filter(model=model)
        
        # Get unique stones (Marquise, Pan, etc.) from raw_stones
        stone_ids = set(raw_stone.stone_type.stone.id for raw_stone in raw_stones)
        stones = Stone.objects.filter(id__in=stone_ids)
        
        # Calculate total weight for the model stones
        all_details = StoneTypeDetail.objects.filter(
            stone_type__in=[rs.stone_type for rs in raw_stones],
            stone__in=stones
        )
        model_total_stone_weight = sum(detail.weight for detail in all_details) or 1  # avoid division by zero
        
        stone_result = []
        
        # Process each stone (Marquise, Pan, etc.)
        for stone in stones:
            stone_types = StoneType.objects.filter(stone=stone, raw_stones__model=model).distinct()
            stone_type_ids = [st.id for st in stone_types]

            # Fetch all stone counts for this stone in this model
            stone_counts = StoneCount.objects.filter(
                model=model,
                stone_type_details__stone_type__in=stone_type_ids,
                stone_type_details__stone=stone
            )

            total_count = stone_counts.aggregate(total=Sum('count'))['total'] or 0
            # Calculate total weight for this stone
            stone_details = StoneTypeDetail.objects.filter(
                stone=stone,
                stone_type__in=stone_types
            )
            stone_total_rate = sum(Decimal(detail.rate) for detail in stone_details)

            stone_total_weight = sum(detail.weight for detail in stone_details) or 0
            
            # Calculate percentage of this stone in the model
            stone_percentage = (stone_total_weight / model_total_stone_weight * 100)


            
            stone_data = {
                'stone_id': stone.id,
                'stone_name': stone.name,
                'total_weight': round(stone_total_weight, 2),  # Add total weight
                'total_rate': round(stone_total_rate, 2),  # Add total rate
                'total_count': total_count,
                'percentage_in_model': round(stone_percentage, 2),
                'stone_distribution': []
            }
            
            # Process each stone type (White, Hydro, etc.) for this stone
            for stone_type in stone_types:
                type_details = StoneTypeDetail.objects.filter(
                    stone=stone,
                    stone_type=stone_type
                )
                
                type_total_weight = sum(detail.weight for detail in type_details) or 0
                
                # Calculate percentage of this type within the stone
                type_percentage_in_stone = (type_total_weight / stone_total_weight * 100) if stone_total_weight > 0 else 0
                
                # Calculate percentage of this type in the overall model
                type_percentage_in_model = (type_total_weight / model_total_stone_weight * 100)

                type_stone_total_rate = sum(detail.rate for detail in type_details) or Decimal(0)

                type_stone_total_weight = sum(detail.weight for detail in type_details) or 0
                stone_type_count = stone_counts.filter(
                    stone_type_details__stone_type=stone_type
                ).aggregate(total=Sum('count'))['total'] or 0

                type_info = {
                    'type_id': stone_type.id,
                    'type_name': stone_type.type_name,
                    'percentage_in_stone': round(type_percentage_in_stone, 2),
                    'percentage_in_model': round(type_percentage_in_model, 2),
                    'type_stone_total_rate':type_stone_total_rate,
                    'type_stone_total_weight':type_stone_total_weight,
                    'count': stone_type_count,
                    'distribution': []
                }
                
                # Add details for this stone type
                for detail in type_details:
                    detail_percentage = (detail.weight / type_total_weight * 100) if type_total_weight > 0 else 0
                    
                    # Get count from StoneCount model
                    detail_count = StoneCount.objects.filter(
                        model=model,
                        stone_type_details=detail
                    ).aggregate(total=Sum('count'))['total'] or 0

                    type_info['distribution'].append({
                        'detail_id': detail.id,
                        'length': detail.length,
                        'breadth': detail.breadth,
                        'weight': str(detail.weight),
                        'rate': str(detail.rate),
                        'percentage': round(detail_percentage, 2),
                        'count': detail_count  # 👈 Add count here
                    })

                
                if type_info['distribution']:
                    stone_data['stone_distribution'].append(type_info)
            
            stone_result.append(stone_data)
            
        # RAW MATERIAL SECTION
        # Get all raw materials used in this model
        raw_materials = RawMaterial.objects.filter(model=model)
        
        # Calculate total weight of all raw materials
        model_total_material_weight = raw_materials.aggregate(total=Sum('weight'))['total'] or 1  # avoid division by zero
        
        material_result = []
        
        # Process each metal
        metal_ids = set(rm.metal.id for rm in raw_materials)
        for metal_id in metal_ids:
            metal = Metal.objects.get(id=metal_id)
            metal_materials = raw_materials.filter(metal=metal)
            
            # Calculate total weight for this metal
            metal_total_weight = metal_materials.aggregate(total=Sum('weight'))['total'] or 0
            
            # Calculate percentage of this metal in the model
            metal_percentage = (metal_total_weight / model_total_material_weight * 100)
            
            material_data = {
                'metal_id': metal.id,
                'metal_unique_id': metal.metal_unique_id,
                'metal_name': metal.name,
                'total_weight': str(metal_total_weight),
                'percentage_in_model': round(metal_percentage, 2)
            }
            
            # Get the latest rate for this metal if exists
            latest_rate = MetalRate.objects.filter(metal=metal).order_by('-date').first()
            if latest_rate:
                material_data.update({
                    'latest_rate': str(latest_rate.rate),
                    'rate_currency': latest_rate.currency,
                    'rate_unit': latest_rate.unit,
                    'rate_weight': str(latest_rate.weight),
                    'rate_date': latest_rate.date.strftime('%Y-%m-%d')
                })
            
            material_result.append(material_data)
        
        return {
            'stone_data': stone_result,
            'material_data': material_result
        }
    
    except Model.DoesNotExist:
        return {'error': 'Model not found'}

def stone_distribution_view(request, model_no):
    stone_data = get_model_distribution(model_no)
    return JsonResponse({'stone_and_material_distribution': stone_data})

# def get_jewelry_types_with_model_count(request):
#     jewelry_data = JewelryType.objects.annotate(model_count=Count('models')).values('id', 'name', 'unique_id', 'model_count')
#     return JsonResponse({'data': list(jewelry_data)})

from collections import defaultdict

def get_models_by_jewelry_type(request, jewelry_type_name=None):
    if jewelry_type_name:
        try:
            jewelry_type = JewelryType.objects.get(name=jewelry_type_name)
        except JewelryType.DoesNotExist:
            try:
                jewelry_id = int(jewelry_type_name)
                jewelry_type = JewelryType.objects.get(id=jewelry_id)
            except (ValueError, JewelryType.DoesNotExist):
                return JsonResponse({'error': 'Jewelry type not found'}, status=404)

        models = Model.objects.filter(jewelry_type=jewelry_type).values(
            'id', 'model_no', 'length', 'breadth', 'weight', 'model_img', 'is_active'
        ).annotate(
            no_of_pieces=Count('model_no'),
            status_name=F('status__status')
        )

        model_ids = [m['id'] for m in models]
        clients_map = (
            ModelClient.objects
            .filter(model_id__in=model_ids)
            .select_related('client')
            .values('model_id', 'client__first_name', 'client__last_name')
        )

        model_clients = defaultdict(list)
        for entry in clients_map:
            full_name = f"{entry['client__first_name']} {entry['client__last_name']}".strip()
            if not full_name:  # fallback in case name is empty
                full_name = 'Unnamed Client'
            model_clients[entry['model_id']].append(full_name)

        for model in models:
            client_list = model_clients.get(model['id'], [])
            model['clients'] = ', '.join(client_list) if client_list else 'N/A'
            

        return JsonResponse({'data': list(models)}, safe=False)
        
# @csrf_exempt
# def create_model(request):
#     if request.method == 'POST':
#         try:
#             # Get form data
#             model_no = request.POST.get('model_no')
#             length = request.POST.get('length')
#             breadth = request.POST.get('breadth')
#             weight = request.POST.get('weight')
#             jewelry_type_id = request.POST.get('jewelry_type')
#             jewelry_type = get_object_or_404(JewelryType, id=jewelry_type_id)
#             model_img = request.FILES.get('model_img')
#             selected_colors = request.POST.getlist('colors[]')
            
#             # Get stones data
#             stones_json = request.POST.get('stones', '[]')
#             stones_data = json.loads(stones_json)
            
#             # Validate required fields
#             if not all([model_no, length, breadth, weight, jewelry_type_id, model_img, selected_colors]):
#                 return JsonResponse({'error': 'All fields are required'}, status=400)

#             # Check if model number already exists
#             if Model.objects.filter(model_no=model_no).exists():
#                 return JsonResponse({'error': 'Model number already exists'}, status=400)

#             # Define the target directory
#             target_directory = os.path.join(settings.BASE_DIR, 'product_inv/static/model_img/')

#             # Ensure the directory exists
#             os.makedirs(target_directory, exist_ok=True)

#             # Extract file extension and rename the file
#             file_extension = os.path.splitext(model_img.name)[1]
#             new_filename = f"{model_no}{file_extension}"
#             file_path = os.path.join(target_directory, new_filename)

#             # Save the file manually
#             with open(file_path, 'wb+') as destination:
#                 for chunk in model_img.chunks():
#                     destination.write(chunk)

#             # Save relative path in the database
#             relative_path = f"model_image/{new_filename}"

#             # Create new model
#             model = Model.objects.create(
#                 model_no=model_no,
#                 length=length,
#                 breadth=breadth,
#                 weight=weight,
#                 model_img=relative_path,
#                 jewelry_type=jewelry_type
#             )
            
#             # Create model colors
#             for color in selected_colors:
#                 ModelColor.objects.create(model=model, color=color)
            
#             # Process stones data
#             for stone_data in stones_data:
#                 # Create RawStones entries
#                 stone_type = get_object_or_404(StoneType, id=stone_data['stone_type_id'])
#                 RawStones.objects.create(
#                     model=model,
#                     stone_type=stone_type
#                 )
                
#                 # Create StoneCount entries if detail_id exists
#                 if 'stone_type_detail_id' in stone_data and stone_data['stone_type_detail_id'] not in [None, '', 'undefined']:
#                     try:
#                         detail_id = int(stone_data['stone_type_detail_id'])  # Convert to int explicitly
#                         stone_type_detail = StoneTypeDetail.objects.get(id=detail_id)
#                         StoneCount.objects.create(
#                             model=model,
#                             stone_type_details=stone_type_detail,
#                             count=stone_data['count']
#                         )
#                     except (ValueError, StoneTypeDetail.DoesNotExist) as e:
#                         print(f"Error creating StoneCount: {e}")
            
#             # Process raw materials data
#             raw_materials_json = request.POST.get('raw_materials', '[]')
#             raw_materials_data = json.loads(raw_materials_json)
            
#             for material_data in raw_materials_data:
#                 metal = get_object_or_404(Metal, id=material_data['material_id'])
#                 RawMaterial.objects.create(
#                     model=model,
#                     metal=metal,
#                     weight=material_data['weight'],
#                     unit='g'  # Default to grams
#                 )
                    
#             return JsonResponse({
#                 'success': True, 
#                 'message': 'Model created successfully',
#                 'model': {
#                     'id': model.id,
#                     'model_no': model.model_no,
#                     'length': float(model.length),
#                     'breadth': float(model.breadth),
#                     'weight': float(model.weight),
#                     'image_path': relative_path,
#                     'color': selected_colors,
#                     'jewelry_type_name': jewelry_type.name
#                 }
#             })
            
#         except Exception as e:
#             import traceback
#             print(traceback.format_exc())
#             return JsonResponse({'error': str(e)}, status=500)

#     return JsonResponse({'error': 'Invalid request method'}, status=405)
# -----------
# @csrf_exempt
# def create_model(request):
#     if request.method == 'POST':
#         try:
#             # Get form data
#             model_no = request.POST.get('model_no')
#             length = request.POST.get('length')
#             breadth = request.POST.get('breadth')
#             weight = request.POST.get('weight')
#             jewelry_type_id = request.POST.get('jewelry_type')
#             jewelry_type = get_object_or_404(JewelryType, id=jewelry_type_id)
#             model_img = request.FILES.get('model_img')
#             selected_colors = request.POST.getlist('colors[]')
#             selected_clients = request.POST.getlist('clients[]')  # Get selected clients
#             status_id = request.POST.get('status')
            
#             model_status = None
#             if status_id:
#                 model_status = get_object_or_404(ModelStatus, id=status_id)
            
#             # Get stones data
#             stones_json = request.POST.get('stones', '[]')
#             stones_data = json.loads(stones_json)

#             if not model_no:
#                 return JsonResponse({'error': 'Model number is required'}, status=400)

            
#             # # Validate required fields
#             # if not all([model_no, length, breadth, weight, jewelry_type_id, model_img, selected_colors,status_id]):
#             #     return JsonResponse({'error': 'All fields are required'}, status=400)

#             # Check if model number already exists
#             if Model.objects.filter(model_no=model_no).exists():
#                 return JsonResponse({'error': 'Model number already exists'}, status=400)

#             # Define the target directory
#             target_directory = os.path.join(settings.BASE_DIR, 'product_inv/static/model_img/')

#             # Ensure the directory exists
#             os.makedirs(target_directory, exist_ok=True)

#             # Extract file extension and rename the file
#             file_extension = os.path.splitext(model_img.name)[1]
#             new_filename = f"{model_no}{file_extension}"
#             file_path = os.path.join(target_directory, new_filename)

#             # Save the file manually
#             with open(file_path, 'wb+') as destination:
#                 for chunk in model_img.chunks():
#                     destination.write(chunk)

#             # Save relative path in the database
#             relative_path = f"model_img/{new_filename}"

#             # Create new model
#             model = Model.objects.create(
#                 model_no=model_no,
#                 length=length,
#                 breadth=breadth,
#                 weight=weight,
#                 model_img=relative_path,
#                 jewelry_type=jewelry_type,
#                 status=model_status
#             )
            
#             # Create model colors
#             for color in selected_colors:
#                 ModelColor.objects.create(model=model, color=color)
            
#             # Create model clients
#             current_user = request.user
#             for client_id in selected_clients:
#                 try:
#                     client_user = User.objects.get(id=client_id)
#                     ModelClient.objects.create(
#                         model=model,
#                         client=client_user,
#                         created_by=current_user,
#                         updated_by=current_user
#                     )
#                 except User.DoesNotExist:
#                     print(f"User with ID {client_id} does not exist")
#                 except Exception as e:
#                     print(f"Error creating ModelClient: {e}")
            
#             # Process stones data
#             for stone_data in stones_data:
#                 # Create RawStones entries
#                 stone_type = get_object_or_404(StoneType, id=stone_data['stone_type_id'])
#                 RawStones.objects.create(
#                     model=model,
#                     stone_type=stone_type
#                 )
                
#                 # Create StoneCount entries if detail_id exists
#                 if 'stone_type_detail_id' in stone_data and stone_data['stone_type_detail_id'] not in [None, '', 'undefined']:
#                     try:
#                         detail_id = int(stone_data['stone_type_detail_id'])  # Convert to int explicitly
#                         stone_type_detail = StoneTypeDetail.objects.get(id=detail_id)
#                         StoneCount.objects.create(
#                             model=model,
#                             stone_type_details=stone_type_detail,
#                             count=stone_data['count']
#                         )
#                     except (ValueError, StoneTypeDetail.DoesNotExist) as e:
#                         print(f"Error creating StoneCount: {e}")
            
#             # Process raw materials data
#             raw_materials_json = request.POST.get('raw_materials', '[]')
#             raw_materials_data = json.loads(raw_materials_json)
            
#             for material_data in raw_materials_data:
#                 metal = get_object_or_404(Metal, id=material_data['material_id'])
#                 RawMaterial.objects.create(
#                     model=model,
#                     metal=metal,
#                     weight=material_data['weight'],
#                     unit='g'  # Default to grams
#                 )
                    
#             return JsonResponse({
#                 'success': True, 
#                 'message': 'Model created successfully',
#                 'model': {
#                     'id': model.id,
#                     'model_no': model.model_no,
#                     'length': float(model.length),
#                     'breadth': float(model.breadth),
#                     'weight': float(model.weight),
#                     'image_path': relative_path,
#                     'color': selected_colors,
#                     'jewelry_type_name': jewelry_type.name,
#                     'status': model_status.status if model_status else None
#                 }
#             })
            
#         except Exception as e:
#             import traceback
#             print(traceback.format_exc())
#             return JsonResponse({'error': str(e)}, status=500)

#     return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
def create_model(request):
    if request.method == 'POST':
        try:
            # Get form data
            model_no = request.POST.get('model_no')
            length = request.POST.get('length')
            breadth = request.POST.get('breadth')
            weight = request.POST.get('weight')
            jewelry_type_id = request.POST.get('jewelry_type')
            model_img = request.FILES.get('model_img')
            selected_colors = request.POST.getlist('colors[]')
            selected_clients = request.POST.getlist('clients[]')
            status_id = request.POST.get('status')
            is_active = request.POST.get('is_active')
            
            # Validate only model_no is required
            if not model_no or model_no.strip() == '':
                return JsonResponse({'error': 'Model number is required'}, status=400)

            # Check if model number already exists
            if Model.objects.filter(model_no=model_no).exists():
                return JsonResponse({'error': 'Model number already exists'}, status=400)

            # Handle jewelry type (optional)
            jewelry_type = None
            if jewelry_type_id:
                jewelry_type = get_object_or_404(JewelryType, id=jewelry_type_id)

            # Handle model status (optional)
            model_status = None
            if status_id:
                model_status = get_object_or_404(ModelStatus, id=status_id)
            
            # Convert empty strings to None for decimal fields
            length = float(length) if length and length.strip() != '' else None
            breadth = float(breadth) if breadth and breadth.strip() != '' else None
            weight = float(weight) if weight and weight.strip() != '' else None
            is_active_value = 'Y' if is_active == 'on' else 'N'
            
            # Handle image upload (optional)
            relative_path = None
            if model_img:
                # Define the target directory
                target_directory = os.path.join(settings.BASE_DIR, 'product_inv/static/model_img/')
                
                # Ensure the directory exists
                os.makedirs(target_directory, exist_ok=True)
                
                # Extract file extension and rename the file
                file_extension = os.path.splitext(model_img.name)[1]
                new_filename = f"{model_no}{file_extension}"
                file_path = os.path.join(target_directory, new_filename)
                
                # Save the file manually
                with open(file_path, 'wb+') as destination:
                    for chunk in model_img.chunks():
                        destination.write(chunk)
                
                # Save relative path in the database
                relative_path = f"model_img/{new_filename}"

            # Create new model with only provided values
            model_data = {
                'model_no': model_no,
                'jewelry_type': jewelry_type,
                'status': model_status,
                'is_active': is_active_value,
                'created_by': None if current_user.is_superuser else current_user,
                'updated_by': None if current_user.is_superuser else current_user
            }
            
            # Add optional fields only if they have values
            if length is not None:
                model_data['length'] = length
            if breadth is not None:
                model_data['breadth'] = breadth
            if weight is not None:
                model_data['weight'] = weight
            if relative_path:
                model_data['model_img'] = relative_path

            model = Model.objects.create(**model_data)
            
            # Create model colors (optional)
            if selected_colors:
                for color in selected_colors:
                    ModelColor.objects.create(model=model, color=color)
            
            # Create model clients (optional)
            if selected_clients:
                current_user = request.user
                for client_id in selected_clients:
                    try:
                        client_user = User.objects.get(id=client_id)
                        ModelClient.objects.create(
                            model=model,
                            client=client_user,
                            created_by=current_user,
                            updated_by=current_user
                        )
                    except User.DoesNotExist:
                        print(f"User with ID {client_id} does not exist")
                    except Exception as e:
                        print(f"Error creating ModelClient: {e}")
            
            # Process stones data (optional)
            stones_json = request.POST.get('stones', '[]')
            if stones_json and stones_json != '[]':
                stones_data = json.loads(stones_json)
                for stone_data in stones_data:
                    # Create RawStones entries
                    stone_type = get_object_or_404(StoneType, id=stone_data['stone_type_id'])
                    RawStones.objects.create(
                        model=model,
                        stone_type=stone_type
                    )
                    
                    # Create StoneCount entries if detail_id exists
                    if 'stone_type_detail_id' in stone_data and stone_data['stone_type_detail_id'] not in [None, '', 'undefined']:
                        try:
                            detail_id = int(stone_data['stone_type_detail_id'])
                            stone_type_detail = StoneTypeDetail.objects.get(id=detail_id)
                            StoneCount.objects.create(
                                model=model,
                                stone_type_details=stone_type_detail,
                                count=stone_data['count']
                            )
                        except (ValueError, StoneTypeDetail.DoesNotExist) as e:
                            print(f"Error creating StoneCount: {e}")
            
            # Process raw materials data (optional)
            raw_materials_json = request.POST.get('raw_materials', '[]')
            if raw_materials_json and raw_materials_json != '[]':
                raw_materials_data = json.loads(raw_materials_json)
                for material_data in raw_materials_data:
                    metal = get_object_or_404(Metal, id=material_data['material_id'])
                    RawMaterial.objects.create(
                        model=model,
                        metal=metal,
                        weight=material_data['weight'],
                        unit='g'
                    )
                    
            return JsonResponse({
                'success': True, 
                'message': 'Model created successfully',
                'model': {
                    'id': model.id,
                    'model_no': model.model_no,
                    'length': float(model.length) if model.length else None,
                    'breadth': float(model.breadth) if model.breadth else None,
                    'weight': float(model.weight) if model.weight else None,
                    'image_path': relative_path,
                    'color': selected_colors,
                    'jewelry_type_name': jewelry_type.name if jewelry_type else None,
                    'status': model_status.status if model_status else None,
                    'is_active': is_active_value
                }
            })
            
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=405)

# New endpoint to get all stones
def get_stones(request):
    stones = Stone.objects.all()
    stone_list = [{'id': stone.id, 'name': stone.name} for stone in stones]
    return JsonResponse(stone_list, safe=False)

# New endpoint to get stone types for a specific stone
def get_stone_types(request, stone_id):
    stone = get_object_or_404(Stone, id=stone_id)
    types = stone.types.all()  # Using the related_name from the model
    type_list = [{'id': t.id, 'type_name': t.type_name} for t in types]
    return JsonResponse(type_list, safe=False)

# # New endpoint to get stone type details
# def get_stone_type_details(request, type_id):
#     try:
#         # Get all details for this type
#         details = StoneTypeDetail.objects.filter(stone_type_id=type_id)
#         if details.exists():
#             details_list = [{
#                 'weight': float(detail.weight),
#                 'length': detail.length,
#                 'breadth': detail.breadth,
#                 'rate': float(detail.rate)
#                 # Add 'shape' back if needed
#                 # 'shape': detail.shape,
#             } for detail in details]
#             return JsonResponse(details_list, safe=False)
#         return JsonResponse([], safe=False)
#     except Exception as e:
#         return JsonResponse({'error': str(e)}, status=500)
    
def get_stone_type_details(request, type_id):
    try:
        # Get all details for this type
        details = StoneTypeDetail.objects.filter(stone_type_id=type_id)
        if details.exists():
            details_list = [{
                'id': detail.id,  # ADD THIS LINE
                'weight': float(detail.weight),
                'length': detail.length,
                'breadth': detail.breadth,
                'rate': float(detail.rate)
            } for detail in details]
            return JsonResponse(details_list, safe=False)
        return JsonResponse([], safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
def get_model(request, model_id):
    model = get_object_or_404(Model, id=model_id)
    model_colors = ModelColor.objects.filter(model=model).values_list('color', flat=True)
    return JsonResponse({
        'id': model.id,
        'model_no': model.model_no,
        'length': float(model.length),
        'breadth': float(model.breadth),
        'weight': float(model.weight),
        'colors': list(model_colors),
    })

@csrf_exempt
def get_materials(request):
    """
    Retrieve all available metals
    """
    if request.method == 'GET':
        metals = Metal.objects.all()
        metal_data = [
            {
                'id': metal.id,
                'name': metal.name,
                'metal_unique_id': metal.metal_unique_id
            } for metal in metals
        ]
        return JsonResponse(metal_data, safe=False)
    return JsonResponse({'error': 'Invalid request method'}, status=405)

@csrf_exempt
def get_material_rate(request, metal_id):
    """
    Get current rate for a specific metal based on weight with enhanced error handling
    """
    if request.method == 'GET':
        try:
            # Get the metal
            try:
                metal = Metal.objects.get(id=metal_id)
            except Metal.DoesNotExist:
                return JsonResponse({'error': f'Metal with ID {metal_id} not found'}, status=404)
            
            # Get the weight from query params
            try:
                weight = float(request.GET.get('weight', 0))
            except ValueError:
                return JsonResponse({'error': 'Invalid weight parameter'}, status=400)
            
            # Get today's rate for the metal
            today = datetime.datetime.now().date()
            
            # Debug logging
            print(f"Searching for metal rate:")
            print(f"Metal ID: {metal_id}")
            print(f"Metal Name: {metal.name}")
            print(f"Weight: {weight} grams")
            print(f"Date: {today}")
            
            # Check if any rates exist for this metal and date
            existing_rates = MetalRate.objects.filter(
                metal=metal, 
                date=today
            )
            
            print(f"Total rates found for this metal and date: {existing_rates.count()}")
            if existing_rates.exists():
                print("Existing rates details:")
                for rate in existing_rates:
                    print(f"- Weight: {rate.weight}, Rate: {rate.rate}")
            
            # First try exact match
            metal_rate = existing_rates.filter(
                weight=weight, 
                unit='gram'
            ).first()
            
            # If no exact match, find closest rate
            if not metal_rate:
                metal_rate = existing_rates.order_by(
                    Abs(F('weight') - weight)
                ).first()
            
            if metal_rate:
                return JsonResponse({
                    'rate': float(metal_rate.rate),
                    'date': metal_rate.date,
                    'currency': metal_rate.currency,
                    'matched_weight': metal_rate.weight  # Added to show which weight was matched
                })
            else:
                return JsonResponse({
                    'error': 'No rate found for this material',
                    'details': {
                        'metal_id': metal_id,
                        'metal_name': metal.name,
                        'requested_weight': weight,
                        'date': today
                    }
                }, status=404)
        
        except Exception as e:
            return JsonResponse({
                'error': 'Unexpected error occurred',
                'details': str(e)
            }, status=500)
    
    return JsonResponse({'error': 'Invalid request method'}, status=405)

import logging
import os

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

@csrf_exempt
def delete_model(request, model_id):
    if request.method == 'DELETE':
        try:
            model = get_object_or_404(Model, id=model_id)

            # Delete associated image
            if model.model_img:  # Ensure model has an image
                image_path = os.path.join(settings.BASE_DIR, 'product_inv/static/', str(model.model_img))
                if os.path.exists(image_path):
                    os.remove(image_path)

            # Delete model entry from database
            model.delete()

            return JsonResponse({'success': True, 'message': 'Model deleted successfully'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request'}, status=400)

@csrf_exempt
def create_jewelry_type(request):
    if request.method == 'POST':
        name = request.POST.get('name')

        if not name:
            return JsonResponse({'success': False, 'error': 'Name is required'})

        if JewelryType.objects.filter(name=name).exists():
            return JsonResponse({'success': False, 'error': 'A jewelry type with this name already exists'})

        unique_id = str(uuid.uuid4())[:8]

        jewelry_type = JewelryType(
            name=name,
            unique_id=unique_id,
        )

        if not request.user.is_superuser:
            jewelry_type.created_by = request.user
            jewelry_type.updated_by = request.user

        jewelry_type.save()

        return JsonResponse({
            'success': True,
            'jewelry_type_id': jewelry_type.id,
            'jewelry_type_name': jewelry_type.name
        })

    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@csrf_exempt
def edit_jewelry_type(request, id):
    if request.method == 'POST':
        name = request.POST.get('name')

        if not name:
            return JsonResponse({'success': False, 'error': 'Name is required'})

        try:
            jewelry_type = JewelryType.objects.get(id=id)
            jewelry_type.name = name  # Removed the comma to avoid tuple assignment

            if not request.user.is_superuser:
                jewelry_type.updated_by = request.user

            jewelry_type.save()
            return JsonResponse({'success': True})
        except JewelryType.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Jewelry Type not found'})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@csrf_exempt
def delete_jewelry_type(request, id):
    if request.method == 'DELETE':
        try:
            jewelry_type = JewelryType.objects.get(id=id)
            jewelry_type.delete()
            return JsonResponse({'success': True})
        except JewelryType.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Jewelry Type not found'})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})

# @csrf_exempt
# def edit_model(request, model_id):
#     if request.method == 'POST':
#         try:
#             model = get_object_or_404(Model, id=model_id)

#             # Get form data
#             model_no = request.POST.get('model_no')
#             length = request.POST.get('length')
#             breadth = request.POST.get('breadth')
#             weight = request.POST.get('weight')
#             jewelry_type_id = request.POST.get('jewelry_type')
#             jewelry_type = get_object_or_404(JewelryType, id=jewelry_type_id)
#             model_img = request.FILES.get('model_img')
#             selected_colors = request.POST.getlist('colors[]')
#             selected_clients = request.POST.getlist('clients[]')  # Get selected clients
#             status_id = request.POST.get('status')

#             model_status = None
#             if status_id:
#                 model_status = get_object_or_404(ModelStatus, id=status_id)

#             stones_json = request.POST.get('stones', '[]')
#             stones_data = json.loads(stones_json)

#             raw_materials_json = request.POST.get('raw_materials', '[]')
#             raw_materials_data = json.loads(raw_materials_json)

#             if not model_no or model_no.strip() == '':
#                 return JsonResponse({'error': 'Model number is required'}, status=400)


#             # Update image if new one is uploaded
#             if model_img:
#                 target_directory = os.path.join(settings.BASE_DIR, 'product_inv/static/model_img/')
#                 os.makedirs(target_directory, exist_ok=True)
#                 file_extension = os.path.splitext(model_img.name)[1]
#                 new_filename = f"{model_no}{file_extension}"
#                 file_path = os.path.join(target_directory, new_filename)

#                 with open(file_path, 'wb+') as destination:
#                     for chunk in model_img.chunks():
#                         destination.write(chunk)

#                 model.model_img = f"model_img/{new_filename}"

#             # Update basic fields
#             model.model_no = model_no
#             model.length = length
#             model.breadth = breadth
#             model.weight = weight
#             model.jewelry_type = jewelry_type
#             model.status = model_status  # Add status update
#             model.save()

#             # Clear and re-create model colors
#             ModelColor.objects.filter(model=model).delete()
#             for color in selected_colors:
#                 ModelColor.objects.create(model=model, color=color)
                
#             # Clear and re-create model clients
#             ModelClient.objects.filter(model=model).delete()
#             current_user = request.user
#             for client_id in selected_clients:
#                 try:
#                     client_user = User.objects.get(id=client_id)
#                     ModelClient.objects.create(
#                         model=model,
#                         client=client_user,
#                         created_by=current_user,
#                         updated_by=current_user
#                     )
#                 except User.DoesNotExist:
#                     print(f"User with ID {client_id} does not exist")
#                 except Exception as e:
#                     print(f"Error creating ModelClient: {e}")

#             # Clear and recreate RawStones and StoneCount
#             RawStones.objects.filter(model=model).delete()
#             StoneCount.objects.filter(model=model).delete()
#             for stone_data in stones_data:
#                 stone_type = get_object_or_404(StoneType, id=stone_data['stone_type_id'])
#                 RawStones.objects.create(model=model, stone_type=stone_type)

#                 if 'stone_type_detail_id' in stone_data and stone_data['stone_type_detail_id'] not in [None, '', 'undefined']:
#                     try:
#                         detail_id = int(stone_data['stone_type_detail_id'])
#                         stone_type_detail = StoneTypeDetail.objects.get(id=detail_id)
#                         StoneCount.objects.create(
#                             model=model,
#                             stone_type_details=stone_type_detail,
#                             count=stone_data['count']
#                         )
#                     except (ValueError, StoneTypeDetail.DoesNotExist):
#                         pass

#             # Clear and recreate RawMaterials
#             RawMaterial.objects.filter(model=model).delete()
#             for material_data in raw_materials_data:
#                 metal = get_object_or_404(Metal, id=material_data['material_id'])
#                 RawMaterial.objects.create(
#                     model=model,
#                     metal=metal,
#                     weight=material_data['weight'],
#                     unit='g'
#                 )

#             return JsonResponse({
#                 'success': True,
#                 'message': 'Model updated successfully',
#                 'model': {
#                     'id': model.id,
#                     'model_no': model.model_no,
#                     'jewelry_type_name': jewelry_type.name,
#                     'status': model_status.status if model_status else None
#                 }
#             })

#         except Exception as e:
#             import traceback
#             print(traceback.format_exc())
#             return JsonResponse({'error': str(e)}, status=500)

#     return JsonResponse({'error': 'Invalid request method'}, status=405)
@csrf_exempt
def edit_model(request, model_id):
    if request.method == 'POST':
        try:
            model = get_object_or_404(Model, id=model_id)

            # Get form data
            model_no = request.POST.get('model_no')
            length = request.POST.get('length')
            breadth = request.POST.get('breadth')
            weight = request.POST.get('weight')
            jewelry_type_id = request.POST.get('jewelry_type')
            jewelry_type = get_object_or_404(JewelryType, id=jewelry_type_id)
            model_img = request.FILES.get('model_img')
            selected_colors = request.POST.getlist('colors[]')
            selected_clients = request.POST.getlist('clients[]')  # Get selected clients
            status_id = request.POST.get('status')
            is_active = request.POST.get('is_active')  # Get is_active value

            model_status = None
            if status_id:
                model_status = get_object_or_404(ModelStatus, id=status_id)

            stones_json = request.POST.get('stones', '[]')
            stones_data = json.loads(stones_json)

            raw_materials_json = request.POST.get('raw_materials', '[]')
            raw_materials_data = json.loads(raw_materials_json)

            if not model_no or model_no.strip() == '':
                return JsonResponse({'error': 'Model number is required'}, status=400)

            # Convert empty strings to None for decimal fields and handle is_active
            length = float(length) if length and length.strip() != '' else None
            breadth = float(breadth) if breadth and breadth.strip() != '' else None
            weight = float(weight) if weight and weight.strip() != '' else None
            is_active_value = 'Y' if is_active == 'on' else 'N'  # Handle is_active like in create_model

            # Update image if new one is uploaded
            if model_img:
                target_directory = os.path.join(settings.BASE_DIR, 'product_inv/static/model_img/')
                os.makedirs(target_directory, exist_ok=True)
                file_extension = os.path.splitext(model_img.name)[1]
                new_filename = f"{model_no}{file_extension}"
                file_path = os.path.join(target_directory, new_filename)

                with open(file_path, 'wb+') as destination:
                    for chunk in model_img.chunks():
                        destination.write(chunk)

                model.model_img = f"model_img/{new_filename}"

            # Update basic fields
            model.model_no = model_no
            model.length = length
            model.breadth = breadth
            model.weight = weight
            model.jewelry_type = jewelry_type
            model.status = model_status  # Add status update
            model.is_active = is_active_value  # Update is_active field
            if not request.user.is_superuser:
                model.updated_by = request.user
            model.save()

            # Clear and re-create model colors
            ModelColor.objects.filter(model=model).delete()
            for color in selected_colors:
                ModelColor.objects.create(model=model, color=color)
                
            # Clear and re-create model clients
            ModelClient.objects.filter(model=model).delete()
            current_user = request.user
            for client_id in selected_clients:
                try:
                    client_user = User.objects.get(id=client_id)
                    ModelClient.objects.create(
                        model=model,
                        client=client_user,
                        created_by=current_user,
                        updated_by=current_user
                    )
                except User.DoesNotExist:
                    print(f"User with ID {client_id} does not exist")
                except Exception as e:
                    print(f"Error creating ModelClient: {e}")

            # Clear and recreate RawStones and StoneCount
            RawStones.objects.filter(model=model).delete()
            StoneCount.objects.filter(model=model).delete()
            for stone_data in stones_data:
                stone_type = get_object_or_404(StoneType, id=stone_data['stone_type_id'])
                RawStones.objects.create(model=model, stone_type=stone_type)

                if 'stone_type_detail_id' in stone_data and stone_data['stone_type_detail_id'] not in [None, '', 'undefined']:
                    try:
                        detail_id = int(stone_data['stone_type_detail_id'])
                        stone_type_detail = StoneTypeDetail.objects.get(id=detail_id)
                        StoneCount.objects.create(
                            model=model,
                            stone_type_details=stone_type_detail,
                            count=stone_data['count']
                        )
                    except (ValueError, StoneTypeDetail.DoesNotExist):
                        pass

            # Clear and recreate RawMaterials
            RawMaterial.objects.filter(model=model).delete()
            for material_data in raw_materials_data:
                metal = get_object_or_404(Metal, id=material_data['material_id'])
                RawMaterial.objects.create(
                    model=model,
                    metal=metal,
                    weight=material_data['weight'],
                    unit='g'
                )

            return JsonResponse({
                'success': True,
                'message': 'Model updated successfully',
                'model': {
                    'id': model.id,
                    'model_no': model.model_no,
                    'length': float(model.length) if model.length else None,
                    'breadth': float(model.breadth) if model.breadth else None,
                    'weight': float(model.weight) if model.weight else None,
                    'jewelry_type_name': jewelry_type.name,
                    'status': model_status.status if model_status else None,
                    'is_active': is_active_value  # Include is_active in response
                }
            })

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=405)

# def get_model_details(request, model_id):
#     model = get_object_or_404(Model, id=model_id)
    
#     # Get model colors
#     selected_colors = [color.color for color in model.model_colors.all()]
    
#     # Get jewelry type
#     jewelry_type = model.jewelry_type
    
#     # Get status
#     model_status = model.status
    
#     # Get image path
#     if model.model_img:
#         relative_path = model.model_img
#         image_path = f"/static/{relative_path}"
#     else:
#         image_path = ""

    
    
#     # Basic model data
#     model_data = {
#         "id": model.id,
#         "model_no": model.model_no,
#         "length": float(model.length),
#         "breadth": float(model.breadth),
#         "weight": float(model.weight),
#         "model_img": image_path,
#         "colors": selected_colors,
#         "jewelry_type_id": jewelry_type.id if jewelry_type else None,
#         "jewelry_type_name": jewelry_type.name if jewelry_type else None,
#         "status_id": model_status.id if model_status else None,
#         "status": model_status.status if model_status else None
#     }
    
#     # --- Clients ---
#     clients = []
#     model_clients = ModelClient.objects.filter(model=model).select_related('client')
#     for client in model_clients:
#         client_data = {
#             "id": client.id,
#             "client_id": client.client.id,
#             "client_name": client.client.get_full_name() or client.client.username,
#             "created_by": client.created_by.username if client.created_by else None,
#             "updated_by": client.updated_by.username if client.updated_by else None,
#             "created_at": client.created_at.strftime("%Y-%m-%d %H:%M:%S") if client.created_at else None,
#             "updated_at": client.updated_at.strftime("%Y-%m-%d %H:%M:%S") if client.updated_at else None,
#         }
#         clients.append(client_data)
    
#     # --- Raw Stones ---
#     # First, enhance raw_stones data by joining with StoneType's related Stone
#     raw_stones = []
#     raw_stones_objects = RawStones.objects.filter(model=model).select_related('stone_type', 'stone_type__stone')
#     for rs in raw_stones_objects:
#         # Get the stone associated with this stone type
#         stone = rs.stone_type.stone if hasattr(rs.stone_type, 'stone') else None
        
#         raw_stone_data = {
#             "id": rs.id,
#             "stone_type_id": rs.stone_type.id,
#             "stone_type_name": rs.stone_type.type_name,
#             "stone_id": stone.id if stone else None,
#             "stone_name": stone.name if stone else "No Stone Selected"
#         }
#         raw_stones.append(raw_stone_data)
    
#     # --- Stone Counts with Complete Details ---
#     stones = []
#     sc_objects = StoneCount.objects.filter(model=model).select_related(
#         'stone_type_details', 
#         'stone_type_details__stone_type', 
#         'stone_type_details__stone'
#     )
    
#     for sc in sc_objects:
#         std = sc.stone_type_details
#         # Make sure we have access to both stone and stone_type data
#         if std:
#             stone = std.stone if hasattr(std, 'stone') else None
#             stone_type = std.stone_type if hasattr(std, 'stone_type') else None
            
#             stone_data = {
#                 "id": sc.id,
#                 "count": sc.count,
#                 "stone_type_detail_id": std.id,
#                 "stone_id": stone.id if stone else None,
#                 "stone_name": stone.name if stone else "No Stone Selected",
#                 "stone_type_id": stone_type.id if stone_type else None,
#                 "stone_type_name": stone_type.type_name if stone_type else "Unknown Type",
#                 "weight": float(std.weight) if hasattr(std, 'weight') and std.weight else 0,
#                 "length": std.length if hasattr(std, 'length') and std.length else 0,
#                 "breadth": std.breadth if hasattr(std, 'breadth') and std.breadth else 0,
#                 "rate": float(std.rate) if hasattr(std, 'rate') and std.rate else 0
#             }
#             stones.append(stone_data)
    
#     # --- Metal Rates Subquery ---
#     latest_rates = MetalRate.objects.filter(
#         metal=OuterRef('metal')
#     ).order_by('-date')  # Gets latest rate per metal
    
#     # --- Raw Materials ---
#     raw_materials = []
#     raw_mats = RawMaterial.objects.filter(model=model).select_related('metal').annotate(
#         latest_rate=Subquery(latest_rates.values('rate')[:1])
#     )
#     for rm in raw_mats:
#         rate = rm.latest_rate or 0
#         material_data = {
#             "id": rm.id,
#             "material_id": rm.metal.id,
#             "material_name": rm.metal.name,
#             "weight": float(rm.weight),
#             "unit": rm.unit,
#             "rate": float(rate),
#             "total_value": float(rm.weight) * float(rate)
#         }
#         raw_materials.append(material_data)
    
#     return JsonResponse({
#         "model": model_data,
#         "clients": clients,
#         "raw_stones": raw_stones,
#         "stones": stones,
#         "raw_materials": raw_materials,
#     })

def get_model_details(request, model_id):
    model = get_object_or_404(Model, id=model_id)
    
    # Get model colors
    selected_colors = [color.color for color in model.model_colors.all()]
    
    # Get jewelry type
    jewelry_type = model.jewelry_type
    
    # Get status
    model_status = model.status
    
    # Get image path
    if model.model_img:
        relative_path = model.model_img
        image_path = f"/static/{relative_path}"
    else:
        image_path = ""

    
    
    # Basic model data - Handle None values for numeric fields
    model_data = {
        "id": model.id,
        "model_no": model.model_no,
        "length": float(model.length) if model.length is not None else 0.0,
        "breadth": float(model.breadth) if model.breadth is not None else 0.0,
        "weight": float(model.weight) if model.weight is not None else 0.0,
        "model_img": image_path,
        "colors": selected_colors,
        "jewelry_type_id": jewelry_type.id if jewelry_type else None,
        "jewelry_type_name": jewelry_type.name if jewelry_type else None,
        "status_id": model_status.id if model_status else None,
        "status": model_status.status if model_status else None
    }
    
    # --- Clients ---
    clients = []
    model_clients = ModelClient.objects.filter(model=model).select_related('client')
    for client in model_clients:
        client_data = {
            "id": client.id,
            "client_id": client.client.id,
            "client_name": client.client.get_full_name() or client.client.username,
            "created_by": client.created_by.username if client.created_by else None,
            "updated_by": client.updated_by.username if client.updated_by else None,
            "created_at": client.created_at.strftime("%Y-%m-%d %H:%M:%S") if client.created_at else None,
            "updated_at": client.updated_at.strftime("%Y-%m-%d %H:%M:%S") if client.updated_at else None,
        }
        clients.append(client_data)
    
    # --- Raw Stones ---
    # First, enhance raw_stones data by joining with StoneType's related Stone
    raw_stones = []
    raw_stones_objects = RawStones.objects.filter(model=model).select_related('stone_type', 'stone_type__stone')
    for rs in raw_stones_objects:
        # Get the stone associated with this stone type
        stone = rs.stone_type.stone if hasattr(rs.stone_type, 'stone') else None
        
        raw_stone_data = {
            "id": rs.id,
            "stone_type_id": rs.stone_type.id,
            "stone_type_name": rs.stone_type.type_name,
            "stone_id": stone.id if stone else None,
            "stone_name": stone.name if stone else "No Stone Selected"
        }
        raw_stones.append(raw_stone_data)
    
    # --- Stone Counts with Complete Details ---
    stones = []
    sc_objects = StoneCount.objects.filter(model=model).select_related(
        'stone_type_details', 
        'stone_type_details__stone_type', 
        'stone_type_details__stone'
    )
    
    for sc in sc_objects:
        std = sc.stone_type_details
        # Make sure we have access to both stone and stone_type data
        if std:
            stone = std.stone if hasattr(std, 'stone') else None
            stone_type = std.stone_type if hasattr(std, 'stone_type') else None
            
            stone_data = {
                "id": sc.id,
                "count": sc.count,
                "stone_type_detail_id": std.id,
                "stone_id": stone.id if stone else None,
                "stone_name": stone.name if stone else "No Stone Selected",
                "stone_type_id": stone_type.id if stone_type else None,
                "stone_type_name": stone_type.type_name if stone_type else "Unknown Type",
                "weight": float(std.weight) if hasattr(std, 'weight') and std.weight is not None else 0.0,
                "length": float(std.length) if hasattr(std, 'length') and std.length is not None else 0.0,
                "breadth": float(std.breadth) if hasattr(std, 'breadth') and std.breadth is not None else 0.0,
                "rate": float(std.rate) if hasattr(std, 'rate') and std.rate is not None else 0.0
            }
            stones.append(stone_data)
    
    # --- Metal Rates Subquery ---
    latest_rates = MetalRate.objects.filter(
        metal=OuterRef('metal')
    ).order_by('-date')  # Gets latest rate per metal
    
    # --- Raw Materials ---
    raw_materials = []
    raw_mats = RawMaterial.objects.filter(model=model).select_related('metal').annotate(
        latest_rate=Subquery(latest_rates.values('rate')[:1])
    )
    for rm in raw_mats:
        rate = rm.latest_rate or 0
        weight = float(rm.weight) if rm.weight is not None else 0.0
        material_data = {
            "id": rm.id,
            "material_id": rm.metal.id,
            "material_name": rm.metal.name,
            "weight": weight,
            "unit": rm.unit,
            "rate": float(rate),
            "total_value": weight * float(rate)
        }
        raw_materials.append(material_data)
    
    return JsonResponse({
        "model": model_data,
        "clients": clients,
        "raw_stones": raw_stones,
        "stones": stones,
        "raw_materials": raw_materials,
    })

def get_clients(request):
    """
    API endpoint to return all clients (users) with client role
    """
    # Get all users with client role
    client_role = Role.objects.filter(role_name='Client').first()
    if client_role:
        client_users = UserRole.objects.filter(role=client_role).select_related('user')
        clients = [{'id': user_role.user.id, 'name': f"{user_role.user.first_name} {user_role.user.last_name}"} for user_role in client_users]
    else:
        # Fallback to get all users if client role doesn't exist
        users = User.objects.all()
        clients = [{'id': user.id, 'name': f"{user.first_name} {user.last_name}"} for user in users]
    
    return JsonResponse(clients, safe=False)


def get_model_clients(request, model_id):
    """
    API endpoint to return all clients associated with a specific model
    """
    try:
        model = get_object_or_404(Model, id=model_id)
        model_clients = ModelClient.objects.filter(model=model).select_related('client')
        clients_list = [{'id': model_client.client.id, 'name': f"{model_client.client.first_name} {model_client.client.last_name}"} 
                       for model_client in model_clients]
        return JsonResponse(clients_list, safe=False)
    except Exception as e:
        print(f"Error fetching model clients: {e}")
        return JsonResponse({'error': str(e)}, status=500)

def get_model_status(request):
    """
    API endpoint to return all model statuses
    """
    statuses = ModelStatus.objects.all()
    statuses_list = [{'id': status.id, 'status': status.status} for status in statuses]
    return JsonResponse(statuses_list, safe=False)

def get_jewelry_types_with_model_count(request):
    # Get all jewelry types with model count using annotation
    # This is more efficient than doing a separate query for each jewelry type
    jewelry_types = JewelryType.objects.annotate(model_count=Count('models'))
    
    data = []
    for jewelry_type in jewelry_types:
        # Get creator and updater information
        created_by = "System"
        updated_by = None
        tracking_info = "No tracking information"
        
        try:
            # Get creator information if available
            if hasattr(jewelry_type, 'created_by') and jewelry_type.created_by:
                created_by_user = jewelry_type.created_by
                created_by = (
                    f"{created_by_user.first_name} {created_by_user.last_name}".strip()
                    if created_by_user else "System"
                )
                if created_by_user and not created_by.strip():
                    created_by = created_by_user.username
            
            # Get updater information if available and different from creator
            if hasattr(jewelry_type, 'updated_by') and jewelry_type.updated_by:
                if not jewelry_type.updated_by or jewelry_type.updated_by == jewelry_type.created_by:
                    updated_by = None
                else:
                    updated_by_user = jewelry_type.updated_by
                    updated_by = f"{updated_by_user.first_name} {updated_by_user.last_name}".strip()
                    if not updated_by.strip():
                        updated_by = updated_by_user.username
            
            # Build tracking info string
            tracking_info = f"Created by {created_by}"
            if updated_by:
                tracking_info += f", Updated by {updated_by}"
                
        except Exception:
            # If any error occurs, use default values
            pass
        
        # Include unique_id if it exists in the jewelry_type model
        item_data = {
            'id': jewelry_type.id,
            'name': jewelry_type.name,
            'model_count': jewelry_type.model_count,
            'created_by': created_by,
            'updated_by': updated_by,
            'tracking_info': tracking_info
        }
        
        # Add unique_id if it exists in the model
        if hasattr(jewelry_type, 'unique_id'):
            item_data['unique_id'] = jewelry_type.unique_id
        
        data.append(item_data)
    
    return JsonResponse({'data': data})
# Add these color options - same as in your HTML
MODEL_COLOR_OPTIONS = [
    'GJ', 'Black', 'White', 'Gray', 'Silver', 'Gold', 'Rose Gold', 'Platinum', 
    'Bronze', 'Copper', 'Red', 'Crimson', 'Maroon', 'Burgundy', 'Pink', 
    'Hot Pink', 'Magenta', 'Salmon', 'Peach', 'Orange', 'Amber', 'Coral', 
    'Yellow', 'Mustard', 'Beige', 'Ivory', 'Lime', 'Olive', 'Green', 
    'Emerald', 'Jade', 'Mint', 'Teal', 'Turquoise', 'Blue', 'Sky Blue', 
    'Azure', 'Cyan', 'Navy Blue', 'Royal Blue', 'Indigo', 'Purple', 
    'Violet', 'Lavender', 'Lilac', 'Brown', 'Chocolate', 'Tan', 'Mahogany', 
    'Bistre', 'Charcoal'
]

def download_sample_model_file(request):
    """Generate and download sample Excel file for bulk model upload"""
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Models"
    
    # Define headers - ADDED CLIENTS COLUMN
    headers = [
        'model_no', 'model_colors', 'clients', 'status', 'length(cm)', 'breadth(cm)', 'weight(gm)'
    ]
    
    # Add headers with styling
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Add sample data (empty rows for user to fill)
    for row_idx in range(2, 102):  # Add 100 empty rows
        ws.cell(row=row_idx, column=1, value='')  # model_no
        ws.cell(row=row_idx, column=2, value='')  # model_colors
        ws.cell(row=row_idx, column=3, value='')  # clients - NEW COLUMN
        ws.cell(row=row_idx, column=4, value='')  # status
        ws.cell(row=row_idx, column=5, value='')  # length
        ws.cell(row=row_idx, column=6, value='')  # breadth
        ws.cell(row=row_idx, column=7, value='')  # weight
    
    # Create data validation for model colors (Column B) - Multiple selection
    color_options_str = ','.join(MODEL_COLOR_OPTIONS)
    colors_validation = DataValidation(
        type="list",
        formula1=f'"{color_options_str}"',
        allow_blank=True
    )
    colors_validation.error = 'Please select from the available colors'
    colors_validation.errorTitle = 'Invalid Color'
    colors_validation.prompt = 'Select multiple colors and separate with comma (e.g., Gold,Silver,Rose Gold)'
    colors_validation.promptTitle = 'Model Colors'
    ws.add_data_validation(colors_validation) 
    colors_validation.add('B2:B101')  # Apply to color column
    
    # UPDATED: Create data validation for clients (Column C) - Fetch using get_clients logic
    try:
        # Get all users with client role (same logic as get_clients function)
        client_role = Role.objects.filter(role_name='Client').first()
        if client_role:
            client_users = UserRole.objects.filter(role=client_role).select_related('user')
            client_options = [f"{user_role.user.first_name} {user_role.user.last_name}" for user_role in client_users]
        else:
            # Fallback to get all users if client role doesn't exist
            users = User.objects.all()
            client_options = [f"{user.first_name} {user.last_name}" for user in users]
        
        if client_options:
            # Limit to reasonable number for dropdown (Excel has limits)
            client_options = client_options[:100]  # Limit to 100 clients
            client_options_str = ','.join([str(client) for client in client_options])
            
            clients_validation = DataValidation(
                type="list",
                formula1=f'"{client_options_str}"',
                allow_blank=True
            )
            clients_validation.error = 'Please select from the available clients'
            clients_validation.errorTitle = 'Invalid Client'
            clients_validation.prompt = 'Select multiple clients and separate with comma (e.g., John Doe,Jane Smith)'
            clients_validation.promptTitle = 'Model Clients'
            ws.add_data_validation(clients_validation)
            clients_validation.add('C2:C101')  # Apply to clients column
                
    except Exception as e:
        print(f"Could not load client options: {e}")
        # Create a sample validation if Role/UserRole models are not available
        sample_clients = "John Doe,Jane Smith,Bob Johnson,Alice Wilson"
        clients_validation = DataValidation(
            type="list",
            formula1=f'"{sample_clients}"',
            allow_blank=True
        )
        clients_validation.error = 'Please select from the available clients'
        clients_validation.errorTitle = 'Invalid Client'
        ws.add_data_validation(clients_validation)
        clients_validation.add('C2:C101')
    
    # Get status options and create validation for status column (Column D - moved from C)
    try:
        status_options = list(ModelStatus.objects.values_list('status', flat=True))
        if status_options:
            status_options_str = ','.join([str(status) for status in status_options])
            status_validation = DataValidation(
                type="list", 
                formula1=f'"{status_options_str}"',
                allow_blank=True
            )
            status_validation.error = 'Please select from the available status options'
            status_validation.errorTitle = 'Invalid Status'
            status_validation.prompt = 'Select one status from the dropdown'
            status_validation.promptTitle = 'Model Status'
            ws.add_data_validation(status_validation)
            status_validation.add('D2:D101')  # Apply to status column (now column D)
    except Exception as e:
        print(f"Could not load status options: {e}")
        # Create a default status validation
        default_status = "Active,Inactive,In Development,Completed"
        status_validation = DataValidation(
            type="list",
            formula1=f'"{default_status}"',
            allow_blank=True
        )
        status_validation.error = 'Please select from the available status options'
        status_validation.errorTitle = 'Invalid Status'
        ws.add_data_validation(status_validation)
        status_validation.add('D2:D101')
    
    # Create Instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        "Instructions for Bulk Model Upload:",
        "",
        "1. MODEL_NO: Required field, must be unique",
        "2. MODEL_COLORS: Click cell to see dropdown with all color options",
        "   - For multiple colors: Select one, then manually add comma and next color",
        "   - Example: Gold,Silver,Rose Gold",
        "3. CLIENTS: Click cell to see dropdown with users having 'Client' role",
        "   - Shows users with Client role (First Name + Last Name format)",
        "   - For multiple clients: Select one, then manually add comma and next client",
        "   - Example: John Doe,Jane Smith,Bob Johnson",
        "   - Use exact full names as shown in dropdown",
        "4. STATUS: Click cell to see dropdown with status options (single selection)",
        "5. LENGTH: Optional, numeric value in cm",
        "6. BREADTH: Optional, numeric value in cm", 
        "7. WEIGHT: Optional, numeric value in gm",
        "",
        "IMPORTANT NOTES:",
        "- Client dropdown shows users with 'Client' role from the system",
        "- Client names are displayed as 'First Name Last Name' format",
        "- If you need to add a new client, ensure they have 'Client' role assigned",
        "- Click on cells in MODEL_COLORS, CLIENTS, and STATUS columns to see dropdown arrows",
        "- For multiple selections: Select first option, then type comma and select next option",
        "- Do not modify the header row",
        "- Empty cells are allowed for optional fields",
        "- Model numbers must be unique in your system",
        "- Client names must match exactly (case-sensitive) as 'First Name Last Name'",
        "",
        "Excel Tips:",
        "- You will see dropdown arrows when you click on color/client/status cells",
        "- For multiple selections: Type 'option1,option2' format",
        "- Dropdowns make data entry faster and reduce errors"
    ]
    
    for row_idx, instruction in enumerate(instructions, 1):
        ws_instructions.cell(row=row_idx, column=1, value=instruction)
        
    # Auto-adjust column width for instructions
    ws_instructions.column_dimensions['A'].width = 80
    
    # Adjust column widths - UPDATED FOR NEW COLUMN
    ws.column_dimensions['A'].width = 15  # model_no
    ws.column_dimensions['B'].width = 30  # model_colors (wider for multiple selections)
    ws.column_dimensions['C'].width = 25  # clients (NEW - wider for multiple selections)
    ws.column_dimensions['D'].width = 20  # status
    ws.column_dimensions['E'].width = 12  # length
    ws.column_dimensions['F'].width = 12  # breadth
    ws.column_dimensions['G'].width = 12  # weight
    
    # Add some sample data in first row to show format - DEFAULT COLORS AND SINGLE ENTRY
    ws.cell(row=2, column=1, value='MODEL001')
    ws.cell(row=2, column=2, value='GJ,White,Gold,Rose Gold')  # Default colors
 
    ws.cell(row=2, column=5, value=2.5)  # length (column moved)
    ws.cell(row=2, column=6, value=1.8)  # breadth (column moved)
    ws.cell(row=2, column=7, value=15.5)  # weight (column moved)
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sample_model_upload.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    return response


@csrf_exempt
def bulk_upload_models(request):
    """Handle bulk upload of models from Excel/CSV file"""
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    
    if 'bulk_file' not in request.FILES:
        return JsonResponse({'error': 'No file uploaded'}, status=400)
    
    file = request.FILES['bulk_file']
    jewelry_type_id = request.POST.get('jewelry_type')
    
    # Validate file extension
    file_ext = os.path.splitext(file.name)[1].lower()
    if file_ext not in ['.xlsx', '.xls', '.csv']:
        return JsonResponse({'error': 'Invalid file format. Please upload Excel or CSV file.'}, status=400)
    
    try:
        # Read file based on extension
        if file_ext == '.csv':
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
        
        # Validate required columns
        required_columns = ['model_no']
        optional_columns = ['model_colors', 'clients', 'status', 'length', 'breadth', 'weight']  # ADDED CLIENTS
        all_expected_columns = required_columns + optional_columns
        
        # Check if required columns exist
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return JsonResponse({
                'error': f'Missing required columns: {", ".join(missing_columns)}'
            }, status=400)
        
        # Clean column names (remove extra spaces)
        df.columns = df.columns.str.strip()
        
        # Process each row
        created_count = 0
        skipped_count = 0
        errors = []
        
        # Get jewelry type if provided
        jewelry_type = None
        if jewelry_type_id:
            try:
                jewelry_type = get_object_or_404(JewelryType, id=jewelry_type_id)
            except:
                jewelry_type = None
        
        # Get current user for client assignments
        current_user = request.user
        
        for index, row in df.iterrows():
            try:
                row_num = index + 2  # +2 because pandas is 0-indexed and we have header row
                
                # Get model_no (required)
                model_no = str(row['model_no']).strip()
                if not model_no or model_no.lower() in ['nan', 'none', '']:
                    errors.append({
                        'row': row_num,
                        'message': 'Model number is required'
                    })
                    continue
                
                # Check if model already exists
                if Model.objects.filter(model_no=model_no).exists():
                    skipped_count += 1
                    errors.append({
                        'row': row_num,
                        'message': f'Model number "{model_no}" already exists - skipped'
                    })
                    continue
                
                # Prepare model data
                model_data = {
                    'model_no': model_no,
                    'jewelry_type': jewelry_type
                }
                
                # Process optional fields
                # Length
                if 'length(cm)' in df.columns and pd.notna(row.get('length(cm)')):
                    try:
                        length_val = float(row['length(cm)'])
                        if length_val > 0:
                            model_data['length'] = length_val
                    except (ValueError, TypeError):
                        errors.append({
                            'row': row_num,
                            'message': f'Invalid length value: {row.get("length")}'
                        })
                        continue
                
                # Breadth
                if 'breadth(cm)' in df.columns and pd.notna(row.get('breadth(cm)')):
                    try:
                        breadth_val = float(row['breadth(cm)'])
                        if breadth_val > 0:
                            model_data['breadth'] = breadth_val
                    except (ValueError, TypeError):
                        errors.append({
                            'row': row_num,
                            'message': f'Invalid breadth value: {row.get("breadth")}'
                        })
                        continue
                
                # Weight
                if 'weight(gm)' in df.columns and pd.notna(row.get('weight(gm)')):
                    try:
                        weight_val = float(row['weight(gm)'])
                        if weight_val > 0:
                            model_data['weight'] = weight_val
                    except (ValueError, TypeError):
                        errors.append({
                            'row': row_num,
                            'message': f'Invalid weight value: {row.get("weight")}'
                        })
                        continue
                
                # Status
                model_status = None
                if 'status' in df.columns and pd.notna(row.get('status')):
                    status_name = str(row['status']).strip()
                    if status_name:
                        try:
                            model_status = ModelStatus.objects.filter(status__iexact=status_name).first()
                            if model_status:
                                model_data['status'] = model_status
                        except:
                            pass  # If ModelStatus model doesn't exist, skip
                
                # Create the model
                model = Model.objects.create(**model_data)
                
                # Process colors
                if 'model_colors' in df.columns and pd.notna(row.get('model_colors')):
                    colors_str = str(row['model_colors']).strip()
                    if colors_str:
                        # Split by comma and clean each color
                        colors = [color.strip() for color in colors_str.split(',') if color.strip()]
                        
                        # Validate colors and create ModelColor entries
                        for color in colors:
                            if color in MODEL_COLOR_OPTIONS:
                                try:
                                    ModelColor.objects.create(model=model, color=color)
                                except:
                                    pass  # If ModelColor model doesn't exist, skip
                            else:
                                errors.append({
                                    'row': row_num,
                                    'message': f'Invalid color "{color}" - model created but color skipped'
                                })
                
                # UPDATED: Process clients - validate using full name format
                if 'clients' in df.columns and pd.notna(row.get('clients')):
                    clients_str = str(row['clients']).strip()
                    if clients_str:
                        # Split by comma and clean each client name
                        client_names = [name.strip() for name in clients_str.split(',') if name.strip()]
                        
                        # Validate clients and create ModelClient entries
                        for client_name in client_names:
                            try:
                                # Split full name to get first and last name
                                name_parts = client_name.split()
                                if len(name_parts) >= 2:
                                    first_name = name_parts[0]
                                    last_name = ' '.join(name_parts[1:])  # Handle multiple last names
                                    
                                    # Find user by first_name and last_name
                                    client_user = User.objects.filter(
                                        first_name__iexact=first_name,
                                        last_name__iexact=last_name,
                                        is_active=True
                                    ).first()
                                    
                                    if client_user:
                                        # Verify user has client role
                                        client_role = Role.objects.filter(role_name='Client').first()
                                        if client_role:
                                            user_has_client_role = UserRole.objects.filter(
                                                user=client_user,
                                                role=client_role
                                            ).exists()
                                            
                                            if not user_has_client_role:
                                                errors.append({
                                                    'row': row_num,
                                                    'message': f'User "{client_name}" does not have Client role - skipped'
                                                })
                                                continue
                                        
                                        # Check if this client-model combination already exists
                                        existing_client = ModelClient.objects.filter(
                                            model=model,
                                            client=client_user
                                        ).first()
                                        
                                        if not existing_client:
                                            ModelClient.objects.create(
                                                model=model,
                                                client=client_user,
                                                created_by=current_user,
                                                updated_by=current_user
                                            )
                                        else:
                                            errors.append({
                                                'row': row_num,
                                                'message': f'Client "{client_name}" already assigned to this model - skipped'
                                            })
                                    else:
                                        errors.append({
                                            'row': row_num,
                                            'message': f'Client "{client_name}" not found - model created but client skipped'
                                        })
                                else:
                                    errors.append({
                                        'row': row_num,
                                        'message': f'Invalid client name format "{client_name}" - use "First Last" format'
                                    })
                                    
                            except Exception as e:
                                errors.append({
                                    'row': row_num,
                                    'message': f'Error adding client "{client_name}": {str(e)}'
                                })
                
                created_count += 1
                
            except Exception as e:
                errors.append({
                    'row': row_num,
                    'message': f'Error processing row: {str(e)}'
                })
                continue
        
        # Return results
        if created_count > 0:
            return JsonResponse({
                'success': True,
                'message': f'Successfully uploaded {created_count} models',
                'created_count': created_count,
                'skipped_count': skipped_count,
                'errors': errors[:10]  # Limit errors shown
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'No models were created',
                'created_count': 0,
                'skipped_count': skipped_count,
                'errors': errors[:20]  # Show more errors if no success
            })
            
    except Exception as e:
        return JsonResponse({
            'error': f'Error processing file: {str(e)}'
        }, status=500)