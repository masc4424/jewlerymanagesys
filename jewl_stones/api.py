from django.http import JsonResponse
from .models import Stone, StoneType, StoneTypeDetail
from product_inv.models import *
from django.db.models import Count
import json
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum
from django.http import JsonResponse
import pandas as pd
import json
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
import logging
from django.http import JsonResponse, HttpResponse
# def get_complete_stone_data(request):
#     stones = Stone.objects.all().prefetch_related('types__details')
#     data = []
    
#     for stone in stones:
#         # Count the total types for this stone
#         type_count = stone.types.count()
        
#         # Add a single entry per stone
#         data.append({
#             'id': stone.id,  # Include the stone ID for edit/delete operations
#             'stone_name': stone.name,
#             'type_count': type_count,
#             'status': stone.is_active  
#         })
            
#     return JsonResponse({'data': data})

def get_complete_stone_data(request):
    stones = Stone.objects.all().prefetch_related('types__details').select_related('created_by', 'updated_by')
    data = []
    
    for stone in stones:
        # Count the total types for this stone
        type_count = stone.types.count()
        
        # Process created_by information
        created_by_user = stone.created_by
        created_by = (
            f"{created_by_user.first_name} {created_by_user.last_name}".strip()
            if created_by_user else "System"
        )
        if created_by_user and not created_by.strip():
            created_by = created_by_user.username
            
        # Process updated_by information (only if different from creator)
        updated_by = None
        if stone.updated_by and stone.updated_by != stone.created_by:
            updated_by_user = stone.updated_by
            updated_by = f"{updated_by_user.first_name} {updated_by_user.last_name}".strip()
            if not updated_by.strip():
                updated_by = updated_by_user.username
                
        # Generate tracking info string
        tracking_info = f"Created by {created_by}"
        if updated_by:
            tracking_info += f", Updated by {updated_by}"
            
        # Add a single entry per stone
        data.append({
            'id': stone.id,
            'stone_name': stone.name,
            'type_count': type_count,
            'status': stone.is_active,
            'created_by': created_by,
            'updated_by': updated_by,
            'tracking_info': tracking_info
        })
        
    return JsonResponse({'data': data}) 

@csrf_exempt
def create_stone(request):
    try:
        stone_name = request.POST.get('stone_name')
        is_active = request.POST.get('is_active')
        
        if not stone_name:
            return JsonResponse({'status': 'error', 'message': 'Stone name is required'})
        
        stone = Stone(
            name=stone_name,
            is_active=is_active
        )

        if not request.user.is_superuser:
            stone.created_by = request.user
            stone.updated_by = request.user

        stone.save()
        
        return JsonResponse({'status': 'success', 'message': 'Stone created successfully'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@csrf_exempt
def update_stone(request):
    try:
        stone_id = request.POST.get('stone_id')
        stone_name = request.POST.get('stone_name')
        is_active = request.POST.get('is_active')
        
        if not stone_id or not stone_name:
            return JsonResponse({'status': 'error', 'message': 'Stone ID and name are required'})
        
        stone = Stone.objects.get(id=stone_id)
        stone.name = stone_name
        stone.is_active = is_active

        if not request.user.is_superuser:
            stone.updated_by = request.user

        stone.save()
        
        return JsonResponse({'status': 'success', 'message': 'Stone updated successfully'})
    except Stone.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Stone not found'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@csrf_exempt
def delete_stone(request):
    try:
        stone_id = request.POST.get('stone_id')
        
        if not stone_id:
            return JsonResponse({'status': 'error', 'message': 'Stone ID is required'})
        
        stone = Stone.objects.get(id=stone_id)
        
        # Check if there are any associated types
        type_count = stone.types.count()
        if type_count > 0:
            return JsonResponse({
                'status': 'error', 
                'message': f'Cannot delete. This stone has {type_count} associated type(s)'
            })
        
        stone.delete()
        
        return JsonResponse({'status': 'success', 'message': 'Stone deleted successfully'})
    except Stone.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Stone not found'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})



def get_model_distribution(model_no):
    try:
        model = Model.objects.get(model_no=model_no)
        
        # STONE SECTION
        # Get all stones used in this model via RawStones
        raw_stones = RawStones.objects.filter(model=model)
        
        # Get unique stones (Marquise, Pan, etc.) from raw_stones
        stone_ids = set(raw_stone.stone_type.stone.id for raw_stone in raw_stones)
        stones = Stone.objects.filter(id__in=stone_ids)
        
        # Calculate total weight for the model stones
        all_details = StoneTypeDetail.objects.filter(
            stone_type__in=[rs.stone_type for rs in raw_stones],
            stone__in=stones
        )
        model_total_stone_weight = sum(detail.weight for detail in all_details) or 1  # avoid division by zero
        
        stone_result = []
        
        # Process each stone (Marquise, Pan, etc.)
        for stone in stones:
            stone_types = StoneType.objects.filter(stone=stone, raw_stones__model=model).distinct()
            
            # Calculate total weight for this stone
            stone_details = StoneTypeDetail.objects.filter(
                stone=stone,
                stone_type__in=stone_types
            )
            stone_total_weight = sum(detail.weight for detail in stone_details) or 0
            
            # Calculate percentage of this stone in the model
            stone_percentage = (stone_total_weight / model_total_stone_weight * 100)
            
            stone_data = {
                'stone_id': stone.id,
                'stone_name': stone.name,
                'percentage_in_model': round(stone_percentage, 2),
                'stone_distribution': []
            }
            
            # Process each stone type (White, Hydro, etc.) for this stone
            for stone_type in stone_types:
                type_details = StoneTypeDetail.objects.filter(
                    stone=stone,
                    stone_type=stone_type
                )
                
                type_total_weight = sum(detail.weight for detail in type_details) or 0
                
                # Calculate percentage of this type within the stone
                type_percentage_in_stone = (type_total_weight / stone_total_weight * 100) if stone_total_weight > 0 else 0
                
                # Calculate percentage of this type in the overall model
                type_percentage_in_model = (type_total_weight / model_total_stone_weight * 100)
                
                type_info = {
                    'type_id': stone_type.id,
                    'type_name': stone_type.type_name,
                    'percentage_in_stone': round(type_percentage_in_stone, 2),
                    'percentage_in_model': round(type_percentage_in_model, 2),
                    'distribution': []
                }
                
                # Add details for this stone type
                for detail in type_details:
                    detail_percentage = (detail.weight / type_total_weight * 100) if type_total_weight > 0 else 0
                    
                    type_info['distribution'].append({
                        'detail_id': detail.id,
                        # 'shape': detail.shape,
                        'length': detail.length,  # Changed from size to length
                        'breadth': detail.breadth,  # Added breadth
                        'weight': str(detail.weight),
                        'rate': str(detail.rate),
                        'percentage': round(detail_percentage, 2)
                    })
                
                stone_data['stone_distribution'].append(type_info)
            
            stone_result.append(stone_data)
            
        # RAW MATERIAL SECTION
        # Get all raw materials used in this model
        raw_materials = RawMaterial.objects.filter(model=model)
        
        # Calculate total weight of all raw materials
        model_total_material_weight = raw_materials.aggregate(total=Sum('weight'))['total'] or 1  # avoid division by zero
        
        material_result = []
        
        # Process each metal
        metal_ids = set(rm.metal.id for rm in raw_materials)
        for metal_id in metal_ids:
            metal = Metal.objects.get(id=metal_id)
            metal_materials = raw_materials.filter(metal=metal)
            
            # Calculate total weight for this metal
            metal_total_weight = metal_materials.aggregate(total=Sum('weight'))['total'] or 0
            
            # Calculate percentage of this metal in the model
            metal_percentage = (metal_total_weight / model_total_material_weight * 100)
            
            material_data = {
                'metal_id': metal.id,
                'metal_unique_id': metal.metal_unique_id,
                'metal_name': metal.name,
                'total_weight': str(metal_total_weight),
                'percentage_in_model': round(metal_percentage, 2)
            }
            
            # Get the latest rate for this metal if exists
            latest_rate = MetalRate.objects.filter(metal=metal).order_by('-date').first()
            if latest_rate:
                material_data.update({
                    'latest_rate': str(latest_rate.rate),
                    'rate_currency': latest_rate.currency,
                    'rate_unit': latest_rate.unit,
                    'rate_weight': str(latest_rate.weight),
                    'rate_date': latest_rate.date.strftime('%Y-%m-%d')
                })
            
            material_result.append(material_data)
        
        return {
            'stone_data': stone_result,
            'material_data': material_result
        }
    
    except Model.DoesNotExist:
        return {'error': 'Model not found'}

def stone_distribution_view(request, model_no):
    stone_data = get_model_distribution(model_no)
    return JsonResponse({'stone_and_material_distribution': stone_data})


# def get_stone_type_data(request):
#     stone_name = request.GET.get('stone_name', '')

#     try:
#         stone = Stone.objects.get(name=stone_name)
#     except Stone.DoesNotExist:
#         return JsonResponse({'data': []})  # Return empty data if stone not found

#     stone_types = StoneType.objects.filter(stone=stone).annotate(detail_count=Count('details'))
#     data = []

#     for stone_type in stone_types:
#         # Add the stone type with detail count
#         data.append({
#             'type_name': stone_type.type_name,
#             'detail_count': stone_type.detail_count,
#         })

#     return JsonResponse({'data': data})

from django.db.models import Count

def get_stone_type_data(request):
    stone_name = request.GET.get('stone_name', '')

    try:
        stone = Stone.objects.get(name=stone_name)
    except Stone.DoesNotExist:
        return JsonResponse({'data': []})  # Return empty data if stone not found

    stone_types = StoneType.objects.filter(stone=stone).annotate(detail_count=Count('details')).select_related('created_by', 'updated_by')
    data = []

    for stone_type in stone_types:
        # Process created_by information
        created_by_user = stone_type.created_by
        created_by = (
            f"{created_by_user.first_name} {created_by_user.last_name}".strip()
            if created_by_user else "System"
        )
        if created_by_user and not created_by.strip():
            created_by = created_by_user.username

        # Process updated_by information (only if different from creator)
        updated_by = None
        if stone_type.updated_by and stone_type.updated_by != stone_type.created_by:
            updated_by_user = stone_type.updated_by
            updated_by = f"{updated_by_user.first_name} {updated_by_user.last_name}".strip()
            if not updated_by.strip():
                updated_by = updated_by_user.username

        # Generate tracking info string
        tracking_info = f"Created by {created_by}"
        if updated_by:
            tracking_info += f", Updated by {updated_by}"

        # Add the stone type with detail count and tracking info
        data.append({
            'type_name': stone_type.type_name,
            'detail_count': stone_type.detail_count,
            'created_by': created_by,
            'updated_by': updated_by,
            'tracking_info': tracking_info,
        })

    return JsonResponse({'data': data})


@csrf_exempt
def create_stone_type(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            stone_name = data.get("stone_name")
            type_name = data.get("type_name")
            

            # Fetch or create the Stone instance
            stone, _ = Stone.objects.get_or_create(name=stone_name)

            # Create the StoneType instance
            stone_type, created = StoneType.objects.get_or_create(
                stone=stone,
                type_name=type_name,
                defaults={
                    'created_by': request.user if not request.user.is_superuser else None,
                    'updated_by': request.user if not request.user.is_superuser else None
                }
            )

            return JsonResponse({"message": "Stone Type created successfully"}, status=201)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"error": "Invalid request method"}, status=405)

@csrf_exempt
def update_stone_type(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            stone_name = data.get("stone_name")
            original_type_name = data.get("original_type_name")
            new_type_name = data.get("new_type_name")
            
            # Find the stone instance
            try:
                stone = Stone.objects.get(name=stone_name)
            except Stone.DoesNotExist:
                return JsonResponse({"error": "Stone not found"}, status=404)
                
            # Find the stone type by stone and original type name
            try:
                stone_type = StoneType.objects.get(stone=stone, type_name=original_type_name)
            except StoneType.DoesNotExist:
                return JsonResponse({"error": "Stone Type not found"}, status=404)
            
            # Update the type name
            stone_type.type_name = new_type_name

            if not request.user.is_superuser:
                stone_type.updated_by = request.user

            stone_type.save()

            
            return JsonResponse({"message": "Stone Type updated successfully"}, status=200)
            
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
            
    return JsonResponse({"error": "Invalid request method"}, status=405)

@csrf_exempt
def delete_stone_type(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            stone_name = data.get("stone_name")
            type_name = data.get("type_name")
            
            # Find the stone instance
            try:
                stone = Stone.objects.get(name=stone_name)
            except Stone.DoesNotExist:
                return JsonResponse({"error": "Stone not found"}, status=404)
                
            # Find the stone type by stone and type name
            try:
                stone_type = StoneType.objects.get(stone=stone, type_name=type_name)
            except StoneType.DoesNotExist:
                return JsonResponse({"error": "Stone Type not found"}, status=404)
            
            # Delete the stone type
            stone_type.delete()
            
            return JsonResponse({"message": "Stone Type deleted successfully"}, status=200)
            
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
            
    return JsonResponse({"error": "Invalid request method"}, status=405)

# def get_stone_type_detail_data(request):
#     stone_name = request.GET.get('stone_name', '')
#     type_name = request.GET.get('type_name', '')

#     try:
#         stone = Stone.objects.get(name=stone_name)
#         stone_type = StoneType.objects.get(stone=stone, type_name=type_name)
#     except (Stone.DoesNotExist, StoneType.DoesNotExist):
#         return JsonResponse({'data': []})  # Return empty data if stone or type not found

#     details = StoneTypeDetail.objects.filter(stone_type=stone_type)
#     data = []

#     for detail in details:
#         data.append({
#             'length': detail.length,
#             'breadth': detail.breadth,
#             'weight': str(detail.weight),
#             'rate': str(detail.rate),
#         })

#     return JsonResponse({'data': data})

# def get_stone_type_detail_data(request):
#     stone_name = request.GET.get('stone_name', '')
#     type_name = request.GET.get('type_name', '')
    
#     print(f"Fetching data for stone_name: {stone_name}, type_name: {type_name}")

#     try:
#         stone = Stone.objects.get(name=stone_name)
#         stone_type = StoneType.objects.get(stone=stone, type_name=type_name)
#     except (Stone.DoesNotExist, StoneType.DoesNotExist) as e:
#         print(f"Error: {str(e)}")
#         return JsonResponse({'data': []})  # Return empty data if stone or type not found

#     details = StoneTypeDetail.objects.filter(stone_type=stone_type)
#     data = []

#     for detail in details:
#         detail_data = {
#             'id': detail.id,  # Make sure this ID is included!
#             'length': detail.length,
#             'breadth': detail.breadth,
#             'weight': str(detail.weight),
#             'rate': str(detail.rate),
#         }
#         print(f"Adding detail: {detail_data}")
#         data.append(detail_data)

#     response_data = {'data': data}
#     print(f"Returning response: {response_data}")
#     return JsonResponse(response_data)

def get_stone_type_detail_data(request):
    stone_name = request.GET.get('stone_name', '')
    type_name = request.GET.get('type_name', '')
    
    print(f"Fetching data for stone_name: {stone_name}, type_name: {type_name}")

    try:
        stone = Stone.objects.get(name=stone_name)
        stone_type = StoneType.objects.get(stone=stone, type_name=type_name)
    except (Stone.DoesNotExist, StoneType.DoesNotExist) as e:
        print(f"Error: {str(e)}")
        return JsonResponse({'data': []})  # Return empty data if stone or type not found

    details = StoneTypeDetail.objects.filter(stone_type=stone_type).select_related('created_by', 'updated_by')
    data = []

    for detail in details:
        # Process created_by information
        created_by_user = detail.created_by
        created_by = (
            f"{created_by_user.first_name} {created_by_user.last_name}".strip()
            if created_by_user else "System"
        )
        if created_by_user and not created_by.strip():
            created_by = created_by_user.username

        # Process updated_by information (only if different from creator)
        updated_by = None
        if detail.updated_by and detail.updated_by != detail.created_by:
            updated_by_user = detail.updated_by
            updated_by = f"{updated_by_user.first_name} {updated_by_user.last_name}".strip()
            if not updated_by.strip():
                updated_by = updated_by_user.username

        # Generate tracking info string
        tracking_info = f"Created by {created_by}"
        if updated_by:
            tracking_info += f", Updated by {updated_by}"

        detail_data = {
            'id': detail.id,
            'length': detail.length,
            'breadth': detail.breadth,
            'weight': str(detail.weight),
            'rate': str(detail.rate),
            'created_by': created_by,
            'updated_by': updated_by,
            'tracking_info': tracking_info,
        }

        print(f"Adding detail: {detail_data}")
        data.append(detail_data)

    response_data = {'data': data}
    print(f"Returning response: {response_data}")
    return JsonResponse(response_data)

@csrf_exempt
def create_stone_type_detail(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            stone_name = data.get("stone_name")
            type_name = data.get("type_name")
            length = data.get("length")
            breadth = data.get("breadth")
            weight = data.get("weight")
            rate = data.get("rate")

            # Get Stone and StoneType
            stone = Stone.objects.get(name=stone_name)
            stone_type = StoneType.objects.get(stone=stone, type_name=type_name)

            # Create the detail
            detail = StoneTypeDetail.objects.create(
                stone=stone,
                stone_type=stone_type,
                length=length,
                breadth=breadth,
                weight=weight,
                rate=rate
            )
            if not request.user.is_superuser:
                detail.created_by = request.user
                detail.updated_by = request.user
                detail.save()

            return JsonResponse({"message": "Detail created successfully"}, status=201)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"error": "Invalid request method"}, status=405)

def get_stone_type_detail(request, detail_id):
    try:
        detail = StoneTypeDetail.objects.get(id=detail_id)
        data = {
            'id': detail.id,
            'length': detail.length,
            'breadth': detail.breadth,
            'weight': str(detail.weight),
            'rate': str(detail.rate),
        }
        return JsonResponse(data)
    except StoneTypeDetail.DoesNotExist:
        return JsonResponse({"error": "Detail not found"}, status=404)
    
@csrf_exempt
def update_stone_type_detail(request, detail_id):
    if request.method == "PUT":
        try:
            data = json.loads(request.body)
            stone_name = data.get("stone_name")
            type_name = data.get("type_name")
            length = data.get("length")
            breadth = data.get("breadth")
            weight = data.get("weight")
            rate = data.get("rate")

            # Get Stone and StoneType
            stone = Stone.objects.get(name=stone_name)
            stone_type = StoneType.objects.get(stone=stone, type_name=type_name)
            
            # Get and update the detail
            detail = StoneTypeDetail.objects.get(id=detail_id)
            detail.length = length
            detail.breadth = breadth
            detail.weight = weight
            detail.rate = rate

            if not request.user.is_superuser:
                detail.updated_by = request.user

            detail.save()

            return JsonResponse({"message": "Detail updated successfully"})

        except StoneTypeDetail.DoesNotExist:
            return JsonResponse({"error": "Detail not found"}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"error": "Invalid request method"}, status=405)

@csrf_exempt
def delete_stone_type_detail(request, detail_id):
    if request.method == "DELETE":
        try:
            detail = StoneTypeDetail.objects.get(id=detail_id)
            detail.delete()
            return JsonResponse({"message": "Detail deleted successfully"})
        except StoneTypeDetail.DoesNotExist:
            return JsonResponse({"error": "Detail not found"}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"error": "Invalid request method"}, status=405)

@csrf_exempt
@login_required
def bulk_upload_stone_type_details(request):
    """Handle bulk upload of stone type details from Excel/CSV files"""
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method"}, status=405)
    
    try:
        # Get parameters
        stone_name = request.POST.get('stone_name')
        type_name = request.POST.get('type_name')
        file = request.FILES.get('bulk_detail_file')
        
        if not all([stone_name, type_name, file]):
            return JsonResponse({
                "success": False,
                "message": "Missing required parameters: stone_name, type_name, or file"
            }, status=400)
        
        # Get Stone and StoneType objects
        try:
            stone = Stone.objects.get(name=stone_name)
            stone_type = StoneType.objects.get(stone=stone, type_name=type_name)
        except Stone.DoesNotExist:
            return JsonResponse({
                "success": False,
                "message": f"Stone '{stone_name}' not found"
            }, status=404)
        except StoneType.DoesNotExist:
            return JsonResponse({
                "success": False,
                "message": f"Stone type '{type_name}' not found for stone '{stone_name}'"
            }, status=404)
        
        # Read the file
        try:
            file_extension = file.name.split('.')[-1].lower()
            
            if file_extension == 'csv':
                df = pd.read_csv(file)
            elif file_extension in ['xlsx', 'xls']:
                df = pd.read_excel(file)
            else:
                return JsonResponse({
                    "success": False,
                    "message": "Unsupported file format. Please use CSV, XLSX, or XLS files."
                }, status=400)
        
        except Exception as e:
            return JsonResponse({
                "success": False,
                "message": f"Error reading file: {str(e)}"
            }, status=400)
        
        # Validate and clean column names
        expected_columns = {
            'length (mm)': 'length',
            'breadth (mm)': 'breadth', 
            'weight (gm)': 'weight',
            'rate (₹)': 'rate'
        }
        
        # Normalize column names (lowercase, strip whitespace)
        df.columns = df.columns.str.lower().str.strip()
        
        # Check for required columns
        missing_columns = []
        for expected_col in expected_columns.keys():
            if expected_col not in df.columns:
                missing_columns.append(expected_col)
        
        if missing_columns:
            return JsonResponse({
                "success": False,
                "message": f"Missing required columns: {', '.join(missing_columns)}"
            }, status=400)
        
        # Rename columns to match model fields
        df = df.rename(columns=expected_columns)
        
        # Remove empty rows
        df = df.dropna(how='all')
        
        # Initialize counters and error tracking
        total_processed = len(df)
        success_count = 0
        error_count = 0
        errors = []
        
        # Process each row
        for index, row in df.iterrows():
            row_number = index + 2  # +2 because index starts at 0 and we have header row
            
            try:
                # Convert fields to float, use 0 if blank/empty
                try:
                    length = float(str(row['length']).strip()) if pd.notna(row['length']) and str(row['length']).strip() != '' else 0.0
                    breadth = float(str(row['breadth']).strip()) if pd.notna(row['breadth']) and str(row['breadth']).strip() != '' else 0.0
                    weight = float(str(row['weight']).strip()) if pd.notna(row['weight']) and str(row['weight']).strip() != '' else 0.0
                    rate = float(str(row['rate']).strip()) if pd.notna(row['rate']) and str(row['rate']).strip() != '' else 0.0
                        
                except (ValueError, TypeError) as e:
                    errors.append({
                        "row": row_number,
                        "message": f"Invalid numeric values: {str(e)}"
                    })
                    error_count += 1
                    continue
                
                # Create the stone type detail (duplicates allowed)
                detail = StoneTypeDetail.objects.create(
                    stone=stone,
                    stone_type=stone_type,
                    length=length,
                    breadth=breadth,
                    weight=weight,
                    rate=rate
                )
                
                # Set created_by and updated_by if user is not superuser
                if not request.user.is_superuser:
                    detail.created_by = request.user
                    detail.updated_by = request.user
                    detail.save()
                
                success_count += 1
                
            except Exception as e:
                errors.append({
                    "row": row_number,
                    "message": f"Unexpected error: {str(e)}"
                })
                error_count += 1
        
        # Prepare response
        if success_count > 0:
            message = f"Successfully processed {success_count} out of {total_processed} records"
            if error_count > 0:
                message += f" ({error_count} errors)"
        else:
            message = f"No records were processed successfully ({error_count} errors)"
        
        return JsonResponse({
            "success": success_count > 0,
            "message": message,
            "total_processed": total_processed,
            "success_count": success_count,
            "error_count": error_count,
            "errors": errors[:50]  # Limit to first 50 errors to avoid large response
        })
        
    except Exception as e:
        return JsonResponse({
            "success": False,
            "message": f"An unexpected error occurred: {str(e)}"
        }, status=500)


@login_required
def download_sample_stone_detail_file(request):
    """Generate and download a sample Excel file for stone type details"""
    try:
        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Stone Type Details Sample"
        
        # Define headers
        headers = ['Length (mm)', 'Breadth (mm)', 'Weight (gm)', 'Rate (₹)']
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Add only one sample data row
        sample_data = [300, 200, 150.5, 1200]
        
        for col_idx, value in enumerate(sample_data, 1):
            ws.cell(row=2, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Create response
        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="stone_type_details_sample.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({
            "error": f"Error generating sample file: {str(e)}"
        }, status=500)